from __future__ import annotations

import json
from pathlib import Path
import sys
import tempfile
import unittest

import pandas as pd
from openpyxl import Workbook, load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src.revamp.constants import LEVEL_COLUMNS, OUTPUT_COLUMNS  # noqa: E402
from src.revamp.classification import classify_po, load_circle_rules  # noqa: E402
from src.revamp.matching import MatchResult, match_sources_to_po  # noqa: E402
from src.revamp.hierarchy import classify_hierarchy_text, load_hierarchy  # noqa: E402
from src.revamp.normalize import canonical_name, normalize_identifier, normalize_npwp  # noqa: E402
from src.revamp.pipeline import _category, build_output_rows, validate_output  # noqa: E402
from src.revamp.source_audit import audit_vendor_sources  # noqa: E402
from src.revamp.readers import read_po  # noqa: E402
from src.revamp.workbook import build_workbook  # noqa: E402


def vendor_record(source: str, **overrides):
    base = {
        "source": source,
        "source_file": f"{source}.xls",
        "source_row": 1,
        "id_vendor": "",
        "sap": "",
        "name": "PT Contoh Abadi",
        "npwp": "",
        "qualification": "",
        "coverage": "",
        "business_field": "",
        "circle": "",
        "status": "",
        "registration_date": "",
        "approval_date": "",
        "email": "",
        "entity_type": "INDIVIDUAL" if source in {"DM", "DM_LAMA", "DCM"} else "COMPANY",
    }
    base.update(overrides)
    return base


def empty_sources():
    return {source: [] for source in ("DRT", "DRT_LAMA", "DM", "DM_LAMA", "DCR", "DCM", "DBCR")}


class RevampPipelineTests(unittest.TestCase):
    def test_normalizers_preserve_identifiers_and_canonicalize_names(self):
        self.assertEqual(normalize_identifier("00123.0"), "00123")
        self.assertEqual(normalize_npwp("01.234.567.8-901.000"), "012345678901000")
        self.assertEqual(canonical_name("PT. Contoh Abadi, Tbk"), "CONTOH ABADI")

    def test_category_precedence_matches_legacy_business_flow(self):
        self.assertEqual(_category(True, True, True, True), "A")
        self.assertEqual(_category(False, True, True, True), "B")
        self.assertEqual(_category(False, False, True, True), "C")
        self.assertEqual(_category(False, False, False, True), "D")
        self.assertEqual(_category(False, False, False, False), "E")

    def test_hierarchy_uses_po_text_and_applies_boundary_rules(self):
        paths, rules = load_hierarchy(PROJECT_ROOT / "config")

        cases = {
            "Pengadaan kabel listrik": ("Supplier", "SUP-11", "Kabel"),
            "Pekerjaan pemasangan kabel listrik": ("Subkontraktor", "SUB-08", "Electrical"),
            "Sewa scaffolding system": ("Alat", "ALT-10", "Scaffolding System"),
            "Pemasangan scaffolding": ("Subkontraktor", "SUB-10", "Scaffolding Erection/Dismantling"),
            "Jasa maintenance excavator": ("Jasa Lainnya", "JAS-02", "Repair"),
            "Sewa bored pile rig": ("Alat", "ALT-06", "Bored Pile Rig"),
            "Electrical Ancillaries": ("Supplier", "SUP-11", "Aksesori Elektrikal"),
        }
        for description, expected in cases.items():
            matches, _ = classify_hierarchy_text(description, paths, rules)
            actual = {(path.level1, path.code, path.level3) for path, _ in matches}
            self.assertIn(expected, actual, description)

        matches, notes = classify_hierarchy_text("Bekisting", paths, rules)
        self.assertEqual(matches, [])
        self.assertTrue(any("AMBIGUOUS" in note for note in notes))
        matches, _ = classify_hierarchy_text("Cut Off Pile", paths, rules)
        self.assertNotIn(
            ("Subkontraktor", "SUB-01", "Cut & Fill"),
            {(path.level1, path.code, path.level3) for path, _ in matches},
        )
        matches, _ = classify_hierarchy_text("Material Alat Bantu", paths, rules)
        self.assertEqual(matches, [])

    def test_po_total_rows_are_reconciled_not_reported_as_missing_vendor(self):
        with tempfile.TemporaryDirectory() as temporary:
            path = Path(temporary) / "PO Test.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Data"
            headers = [
                "Doc.Date", "PO", "Item.PO", "Divisi", "Nama Divisi", "Vendor",
                "Nama Vendor", "Material", "Deskripsi", "Project/KP", "Qty", "Unit",
                "Nilai PO", "Harga Satuan", "Currency",
            ]
            sheet.append([])
            sheet.append([])
            sheet.append(headers)
            sheet.append(["", "4501", "10", "", "", "100", "Vendor", "M1", "Semen", "P1", "1", "EA", "1000", "10", "IDR"])
            sheet.append(["", "", "", "", "", "", "", "", "", "", "", "", "1000", "10", "IDR"])
            workbook.save(path)

            po, stats = read_po(path, "HK")

            self.assertEqual(len(po), 1)
            self.assertEqual(stats["blank_vendor_rows"], 0)
            self.assertEqual(stats["footer_rows"], 1)
            self.assertEqual(stats["footer_reconciliations"][0]["Status Nilai PO"], "SESUAI")

    def test_candidate_id_links_to_master_sap_before_name(self):
        po = pd.DataFrame([{"sap": "2020000001", "name": "Vendor PO"}])
        sources = empty_sources()
        sources["DM"].append(
            vendor_record("DM", id_vendor="8000001", sap="2020000001", name="Nama Master")
        )
        sources["DCM"].append(
            vendor_record("DCM", id_vendor="8000001", name="Nama Calon Berbeda")
        )
        result = match_sources_to_po(po, sources)
        self.assertEqual(
            result.matched["2020000001"]["DCM"][0]["match_method"], "ID_LINK"
        )

    def test_ambiguous_name_is_not_auto_merged(self):
        po = pd.DataFrame(
            [
                {"sap": "1", "name": "PT Nama Sama"},
                {"sap": "2", "name": "PT Nama Sama"},
            ]
        )
        sources = empty_sources()
        sources["DBCR"].append(vendor_record("DBCR", name="PT Nama Sama"))
        result = match_sources_to_po(po, sources)
        self.assertNotIn("1", result.matched)
        self.assertNotIn("2", result.matched)
        self.assertEqual(result.review_rows[0]["Issue"], "AMBIGUOUS_EXACT_NAME")

    def test_all_source_rows_outside_po_are_reported_with_line_number(self):
        po = pd.DataFrame([{"sap": "1", "name": "Vendor PO"}])
        sources = empty_sources()
        sources["DRT"].append(
            vendor_record("DRT", sap="9", name="Vendor Non PO", source_row=27)
        )
        result = match_sources_to_po(po, sources)
        review = next(row for row in result.review_rows if row["Issue"] == "SOURCE_SAP_NOT_IN_PO")
        self.assertEqual(review["Source Row"], "27")
        self.assertEqual(review["NO SAP"], "9")

    def test_source_audit_reports_duplicates_and_missing_required_data(self):
        sources = empty_sources()
        sources["DRT"] = [
            vendor_record("DRT", sap="100", id_vendor="10", name="PT Sama", source_row=5),
            vendor_record("DRT", sap="100", id_vendor="10", name="PT Sama", source_row=9),
            vendor_record("DRT", sap="", id_vendor="", name="PT Kosong", source_row=12),
        ]
        stats = {
            source: {"rejected_rows": []}
            for source in sources
        }
        reviews = audit_vendor_sources(sources, stats)
        by_issue = {row["Issue"]: row for row in reviews}
        self.assertEqual(by_issue["SOURCE_DUPLICATE_SAP"]["Source Row"], "5, 9")
        self.assertEqual(by_issue["SOURCE_DUPLICATE_ID"]["Severity"], "HIGH")
        self.assertEqual(by_issue["SOURCE_RECORD_MISSING_SAP"]["Source Row"], "12")

    def test_output_keeps_level_columns_and_balance_blank(self):
        po = pd.DataFrame([{"sap": "1"}])
        classified = {
            "1": {
                "names": ["Vendor Satu"],
                "po_sources": {"HK"},
                "item_text": "Pekerjaan contoh",
                "item_text_truncated": False,
                "final_classification": "PEKERJAAN SIPIL STRUCTURE",
            }
        }
        matches = MatchResult(matched={}, review_rows=[], method_counts={}, outside_po_counts={})
        settings = {
            "po_only_entity_type": "Perusahaan",
            "categories": {
                "A": {"treatment": "DRP"},
                "B": {"treatment": "DRP + Daftar ulang"},
                "C": {"treatment": "DRP + Prioritas Approve"},
                "D": {"treatment": "DRP + Update data"},
                "E": {"treatment": "DRP + Daftar ulang"},
            },
        }
        rows, _ = build_output_rows(po, classified, matches, settings)
        validate_output(rows, 1)
        self.assertEqual(rows[0]["Kategori"], "E")
        self.assertEqual(rows[0]["Inject"], "✓")
        self.assertTrue(all(rows[0][column] == "" for column in LEVEL_COLUMNS))
        self.assertEqual(rows[0]["Saldo Hutang"], "")

    def test_output_reports_missing_identity_tax_and_classification(self):
        po = pd.DataFrame([{"sap": "1"}])
        classified = {
            "1": {
                "names": ["Vendor Belum Lengkap"],
                "po_sources": {"JO"},
                "item_text": "Uraian yang belum memiliki rule",
                "item_text_truncated": False,
                "final_classification": "",
            }
        }
        matches = MatchResult(matched={}, review_rows=[], method_counts={}, outside_po_counts={})
        settings = {
            "po_only_entity_type": "Perusahaan",
            "categories": {
                "A": {"treatment": "DRP"},
                "B": {"treatment": "DRP + Daftar ulang"},
                "C": {"treatment": "DRP + Prioritas Approve"},
                "D": {"treatment": "DRP + Update data"},
                "E": {"treatment": "DRP + Daftar ulang"},
            },
        }

        rows, reviews = build_output_rows(po, classified, matches, settings)

        issues = {review["Issue"] for review in reviews}
        self.assertEqual(rows[0]["NO SAP"], "1")
        self.assertIn("PO_VENDOR_NO_REGISTRY_MATCH", issues)
        self.assertIn("MISSING_ID_VENDOR", issues)
        self.assertIn("MISSING_NPWP", issues)
        self.assertIn("PO_RULE_GAP_CIRCLE_EMPTY", issues)

    def test_po_classification_is_specific_and_generic_text_stays_unresolved(self):
        po = pd.DataFrame(
            [
                {
                    "company": "HK", "po": "1", "item_po": "10", "sap": "100",
                    "name": "Vendor MEP", "description": "Pekerjaan MEP gedung",
                    "material": "", "division": "", "project": "P1",
                },
                {
                    "company": "HK", "po": "2", "item_po": "10", "sap": "200",
                    "name": "Vendor Pintu", "description": "Pengadaan fire rated steel door",
                    "material": "", "division": "", "project": "P1",
                },
                {
                    "company": "JO", "po": "3", "item_po": "10", "sap": "300",
                    "name": "Vendor Umum", "description": "Material alat bantu",
                    "material": "", "division": "", "project": "P2",
                },
                {
                    "company": "JO", "po": "4", "item_po": "10", "sap": "400",
                    "name": "Vendor Perancah", "description": "Sewa scafolding",
                    "material": "", "division": "", "project": "P2",
                },
                {
                    "company": "JO", "po": "5", "item_po": "10", "sap": "500",
                    "name": "Vendor Langsir", "description": "Upah langsir besi",
                    "material": "", "division": "", "project": "P2",
                },
                {
                    "company": "JO", "po": "6", "item_po": "10", "sap": "600",
                    "name": "Vendor Upah", "description": "Upah",
                    "material": "", "division": "", "project": "P2",
                },
            ]
        )

        classified, _, unresolved = classify_po(po, PROJECT_ROOT / "config")

        self.assertEqual(classified["100"]["ordered_labels"], ["MECHANICAL DAN ELECTRICAL"])
        self.assertEqual(classified["200"]["ordered_labels"], ["PINTU TAHAN API"])
        self.assertEqual(classified["300"]["ordered_labels"], [])
        self.assertEqual(classified["400"]["ordered_labels"], ["SCAFFOLDING / PERANCAH"])
        self.assertEqual(classified["500"]["ordered_labels"], ["LOGISTICS / TRANSPORT"])
        self.assertEqual(classified["600"]["ordered_labels"], [])
        self.assertEqual(len(unresolved), 2)

    def test_circle_validates_but_does_not_add_unrelated_final_classification(self):
        po = pd.DataFrame([{"sap": "1"}])
        classified = {
            "1": {
                "names": ["Vendor Satu"],
                "po_sources": {"HK"},
                "item_text": "Sewa vibro roller",
                "item_text_truncated": False,
                "ordered_labels": ["EQUIPMENT / RENTAL"],
                "final_classification": "EQUIPMENT / RENTAL",
            }
        }
        drt = vendor_record(
            "DRT", sap="1", name="Vendor Satu", circle="KONSTRUKSI BAJA"
        )
        matches = MatchResult(
            matched={"1": {"DRT": [drt]}},
            review_rows=[],
            method_counts={},
            outside_po_counts={},
        )
        rows, reviews = build_output_rows(
            po,
            classified,
            matches,
            {
                "po_only_entity_type": "Perusahaan",
                "categories": {
                    "A": {"treatment": "DRP"},
                    "B": {"treatment": "DRP + Daftar ulang"},
                    "C": {"treatment": "DRP + Prioritas Approve"},
                    "D": {"treatment": "DRP + Update data"},
                    "E": {"treatment": "DRP + Daftar ulang"},
                },
            },
            circle_rules=load_circle_rules(PROJECT_ROOT / "config"),
        )

        self.assertEqual(rows[0]["Klasifikasi Final"], "EQUIPMENT / RENTAL")
        self.assertNotIn("STEEL / FABRIKASI", rows[0]["Klasifikasi Final"])
        self.assertIn("PO_CIRCLE_NO_OVERLAP", {row["Issue"] for row in reviews})

    def test_output_guard_compares_exact_po_vendor_universe(self):
        row = {column: "" for column in OUTPUT_COLUMNS}
        row.update({"NO SAP": "1", "PO": "✓"})
        with self.assertRaises(RuntimeError):
            validate_output([row], {"2"})

    def test_workbook_builder_creates_two_sheet_auditable_output(self):
        data_row = {column: "" for column in OUTPUT_COLUMNS}
        data_row.update(
            {
                "ID Vendor": "00123",
                "NO SAP": "2020000001",
                "Nama Rekanan": "PT Contoh",
                "NPWP": "012345678901000",
                "PO": "✓",
                "Inject": "✓",
                "Kategori": "E",
                "Perlakuan": "DRP + Daftar ulang",
                LEVEL_COLUMNS[0]: "Supplier",
                LEVEL_COLUMNS[1]: "SUP-01 | Material Semen, Beton & Grout",
                LEVEL_COLUMNS[2]: "SUP-01 | Semen",
            }
        )
        bundle = {
            "output_columns": OUTPUT_COLUMNS,
            "data_rows": [data_row],
            "summary_rows": [
                {
                    "Metrik": "Vendor output unik",
                    "Nilai": 1,
                    "Keterangan": "Satu baris per NO SAP",
                }
            ],
            "review_rows": [
                {
                    "Severity": "HIGH",
                    "Issue": "PO_VENDOR_NO_REGISTRY_MATCH",
                    "Source": "PO HK",
                    "Source Row": "",
                    "ID Vendor": "00123",
                    "NO SAP": "2020000001",
                    "Nama Rekanan": "PT Contoh",
                    "Match Method": "PO_ONLY",
                    "Detail": "Vendor belum ditemukan pada registry.",
                },
                {
                    "Severity": "HIGH", "Issue": "SOURCE_DUPLICATE_SAP",
                    "Source": "DRT", "Source Row": "5, 9", "ID Vendor": "00123",
                    "NO SAP": "2020000001", "Nama Rekanan": "PT Contoh",
                    "Match Method": "WITHIN_SOURCE_DUPLICATE", "Detail": "Duplikasi SAP.",
                },
                {
                    "Severity": "MEDIUM", "Issue": "MISSING_ID_VENDOR",
                    "Source": "PO HK", "Source Row": "", "ID Vendor": "",
                    "NO SAP": "2020000001", "Nama Rekanan": "PT Contoh",
                    "Match Method": "OUTPUT_COMPLETENESS", "Detail": "ID belum tersedia.",
                },
                {
                    "Severity": "MEDIUM", "Issue": "PO_RULE_GAP_CIRCLE_EMPTY",
                    "Source": "PO HK", "Source Row": "", "ID Vendor": "00123",
                    "NO SAP": "2020000001", "Nama Rekanan": "PT Contoh",
                    "Match Method": "NO_PO_RULE_MATCH", "Detail": "Belum ada rule.",
                },
            ],
            "evidence_rows": [
                {
                    "NO SAP": "2020000001",
                    "Nama Vendor PO": "PT Contoh",
                    "Rank": 1,
                    "Klasifikasi": "EQUIPMENT / RENTAL",
                    "Jumlah PO Berbeda": 1,
                    "Jumlah Item PO": 1,
                    "Rule ID": "PO-001",
                    "Confidence Rule": "HIGH",
                    "Dukungan Circle": "CIRCLE KOSONG",
                    "Sumber Final": "PO",
                    "Contoh Deskripsi": "Sewa alat",
                    "Rule Pattern": "SEWA ALAT",
                }
            ],
            "unresolved_rows": [
                {
                    "Company": "HK",
                    "NO SAP": "2020000001",
                    "Nama Vendor": "PT Contoh",
                    "Deskripsi Belum Terklasifikasi": "Material bantu",
                    "Jumlah Item": 1,
                    "Contoh PO": "450000001",
                    "Contoh Item PO": "10",
                    "Contoh Project": "P1",
                    "Tindakan": "Review",
                }
            ],
            "hierarchy_evidence_rows": [
                {
                    "NO SAP": "2020000001", "Rank": 1, "Level 1": "Supplier",
                    "Kode Level 2": "SUP-01", "Level 2": "Material Semen, Beton & Grout",
                    "Level 3": "Semen", "Jumlah PO Berbeda": 1, "Jumlah Item PO": 1,
                    "Baris Sumber PO": "HK:4", "Bukti Istilah": "Semen",
                    "Contoh Deskripsi": "Pengadaan semen", "Boundary Diterapkan": "",
                }
            ],
            "hierarchy_unresolved_rows": [
                {
                    "Company": "HK", "NO SAP": "2020000001", "Nama Vendor": "PT Contoh",
                    "Deskripsi Belum Memiliki Level": "Material bantu", "Alasan": "NO_MATCH",
                    "Detail": "", "Jumlah Item": 1, "Contoh PO": "450000001",
                    "Contoh Item PO": "20", "Baris Sumber PO": "HK:5",
                    "Contoh Project": "P1", "Tindakan": "Review",
                }
            ],
            "po_footer_rows": [
                {
                    "Company": "HK", "Source File": "PO HK.xlsx", "Source Row": "10",
                    "Currency": "IDR", "Nilai PO Footer": "1000",
                    "Nilai PO Hitung Ulang": "1000", "Status Nilai PO": "SESUAI",
                    "Harga Satuan Footer": "10", "Harga Satuan Hitung Ulang": "10",
                    "Status Harga Satuan": "SESUAI", "Keterangan": "Baris total/footer.",
                }
            ],
            "assumptions": ["PO HK dan PO JO menjadi universe output."],
        }

        with tempfile.TemporaryDirectory() as temporary:
            bundle_path = Path(temporary) / "bundle.json"
            output_path = Path(temporary) / "output.xlsx"
            bundle_path.write_text(json.dumps(bundle), encoding="utf-8")
            build_workbook(bundle_path, output_path)

            workbook = load_workbook(output_path, data_only=False)
            self.assertEqual(workbook.sheetnames, ["Data Cleansing", "Audit"])
            self.assertEqual(workbook["Data Cleansing"]["A2"].value, "00123")
            self.assertEqual(workbook["Data Cleansing"]["D2"].value, "012345678901000")
            self.assertEqual(workbook["Data Cleansing"]["A2"].number_format, "@")
            self.assertIn("DataCleansingTable", workbook["Data Cleansing"].tables)
            self.assertIn("AuditDuplicatesTable", workbook["Audit"].tables)
            self.assertIn("AuditCompletenessTable", workbook["Audit"].tables)
            self.assertIn("AuditMatchingTable", workbook["Audit"].tables)
            self.assertIn("AuditClassificationReviewTable", workbook["Audit"].tables)
            self.assertIn("AuditClassificationEvidenceTable", workbook["Audit"].tables)
            self.assertNotIn("AuditPOFooterTable", workbook["Audit"].tables)
            self.assertIn("AuditHierarchyEvidenceTable", workbook["Audit"].tables)
            self.assertIn("AuditHierarchyUnresolvedTable", workbook["Audit"].tables)
            self.assertIn("AuditUnresolvedPOTable", workbook["Audit"].tables)
            self.assertEqual(workbook["Data Cleansing"]["B2"].fill.fgColor.rgb, "00F4CCCC")
            self.assertEqual(workbook["Data Cleansing"]["G2"].fill.fgColor.rgb, "00F4CCCC")


class CircleRecordTests(unittest.TestCase):
    def test_equipment_model_does_not_imply_computer_or_unspecified_subtype(self):
        paths, rules = load_hierarchy(PROJECT_ROOT / 'config')
        matches, _ = classify_hierarchy_text('RO Excavator PC 200', paths, rules)
        self.assertEqual({(p.level1, p.level3) for p, _ in matches}, {('Alat', 'Excavator')})
        for text, forbidden in [('Sewa crane', 'Mobile Crane'), ('Sewa loader', 'Wheel Loader')]:
            matches, _ = classify_hierarchy_text(text, paths, rules)
            self.assertNotIn(forbidden, {p.level3 for p, _ in matches})

    def test_compact_bundle_roundtrip_reuses_long_text(self):
        from src.revamp.bundle import read_bundle, write_bundle
        text = 'Pengadaan semen; ' * 1000
        payload = {'data_rows': [{'Item': text}, {'Item': text}], 'value': 100.25}
        with tempfile.TemporaryDirectory() as folder:
            path = Path(folder) / 'bundle.json'
            write_bundle(path, payload)
            restored = read_bundle(path)
            self.assertEqual(restored, payload)
            self.assertIs(restored['data_rows'][0]['Item'], restored['data_rows'][1]['Item'])

    def test_unposted_sap_is_missing_not_a_shared_identity(self):
        from src.revamp.readers import normalize_sap
        self.assertEqual(normalize_sap('not_posted'), '')
        self.assertEqual(normalize_sap('NOT_POSTED'), '')
        self.assertEqual(normalize_sap('001230'), '001230')

    def assemble(self, sources, items):
        from src.revamp.circle_output import assemble_circle_output
        from src.revamp.hierarchy import classify_po_hierarchy
        po = pd.DataFrame([dict(company='HK', source_file='PO HK.xlsx', source_row=str(i + 4),
            doc_date='', po=str(i + 1), item_po='10', sap='100', name='PT Sama',
            material='', description='Pengadaan semen', division='', project='', po_value='100',
            unit_price='9999', currency='IDR', **{}) | item for i, item in enumerate(items)])
        classified, _, _ = classify_po(po, PROJECT_ROOT / 'config')
        hierarchy, _, _, _ = classify_po_hierarchy(po, PROJECT_ROOT / 'config')
        settings = json.loads((PROJECT_ROOT / 'config/revamp_settings.json').read_text())
        return assemble_circle_output(po, sources, match_sources_to_po(po, sources), classified, hierarchy, settings)

    def test_all_circle_records_retained_and_saldo_once_per_sap(self):
        sources = empty_sources()
        sources['DRT'] = [vendor_record('DRT', sap='100', source_row=3), vendor_record('DRT', sap='100', source_row=4)]
        sources['DBCR'] = [vendor_record('DBCR', sap='900', source_row=2)]
        rows, reviews, origins, ledger = self.assemble(sources, [{'po_value': '12.25'}, {'po_value': '20.75', 'company': 'JO'}])
        self.assertEqual(len(rows), 3)
        self.assertEqual([r['Saldo Hutang'] for r in rows], [33.0, '', ''])
        self.assertEqual(origins[1]['Baris Pemilik Saldo'], 2)
        self.assertEqual(rows[2]['PO'], '')
        self.assertEqual(rows[2]['Klasifikasi Final'], '')
        self.assertEqual(ledger[0]['Jumlah Item PO'], 2)

    def test_same_name_different_sap_is_flagged_not_merged(self):
        sources = empty_sources()
        sources['DCR'] = [vendor_record('DCR', sap='100', name='PT Sama', source_row=3),
                          vendor_record('DCR', sap='200', name='PT Sama', source_row=4)]
        rows, reviews, _, _ = self.assemble(sources, [{}, {'sap': '200', 'po_value': '300'}])
        self.assertEqual([r['Saldo Hutang'] for r in rows], [100.0, 300.0])
        self.assertEqual(sum(r['Issue'] == 'NAME_MULTIPLE_SAP' for r in reviews), 2)
        self.assertTrue(all(r['Klasifikasi Final'] for r in rows))

    def test_name_only_candidate_keeps_missing_sap_and_is_classified(self):
        sources = empty_sources()
        sources['DBCR'] = [vendor_record('DBCR', name='PT Sama', source_row=2)]
        rows, reviews, _, _ = self.assemble(sources, [{}])
        self.assertEqual(rows[0]['NO SAP'], '')
        self.assertEqual(rows[0]['Saldo Hutang'], '')
        self.assertTrue(rows[0]['Klasifikasi Final'])
        self.assertEqual(rows[0][LEVEL_COLUMNS[2]], 'Semen')
        self.assertEqual(rows[1]['Saldo Hutang'], 100.0)
        self.assertIn('NAME_LINK_REQUIRES_CONFIRMATION', {r['Issue'] for r in reviews})

    def test_missing_name_record_is_not_dropped(self):
        sources = empty_sources()
        sources['DCM'] = [vendor_record('DCM', sap='100', name='', source_row=3)]
        rows, _, _, _ = self.assemble(sources, [{}])
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]['Nama Rekanan'], '')
        self.assertEqual(rows[0]['Saldo Hutang'], 100.0)

    def test_repeated_po_rows_are_both_counted_and_currency_override_visible(self):
        rows, reviews, _, _ = self.assemble(empty_sources(), [{'po': '1', 'currency': 'USD'}, {'po': '1', 'currency': 'USD'}])
        self.assertEqual(rows[0]['Saldo Hutang'], 200.0)
        self.assertIn('CURRENCY_LABEL_OVERRIDE', {r['Issue'] for r in reviews})

    def test_invalid_amount_does_not_become_zero(self):
        for value in ['', 'abc', 'NaN', 'Infinity']:
            with self.subTest(value=value), self.assertRaises(ValueError):
                self.assemble(empty_sources(), [{'po_value': value}])

    def test_scopes_are_only_po_evidenced_without_codes(self):
        rows, _, _, _ = self.assemble(empty_sources(), [{}, {'description': 'Pengadaan grout'}])
        self.assertEqual(rows[0][LEVEL_COLUMNS[0]], 'Supplier')
        self.assertEqual(rows[0][LEVEL_COLUMNS[1]], 'Material Semen, Beton & Grout')
        self.assertEqual(set(rows[0][LEVEL_COLUMNS[2]].split('; ')), {'Semen', 'Grout'})
        self.assertNotIn('Mortar', rows[0][LEVEL_COLUMNS[2]])

    def test_highlight_references_all_duplicate_rows_and_missing_sap(self):
        from src.revamp.circle_output import attach_output_references
        sources = empty_sources()
        sources['DRT'] = [vendor_record('DRT', sap='100', source_row=3), vendor_record('DRT', sap='100', source_row=4)]
        sources['DCM'] = [vendor_record('DCM', name='Unknown', source_row=3)]
        rows, reviews, origins, ledger = self.assemble(sources, [{}])
        reviews.extend(audit_vendor_sources(sources, {}))
        reviews = attach_output_references(reviews, origins)
        dup = next(r for r in reviews if r['Issue'] == 'SOURCE_DUPLICATE_SAP')
        self.assertEqual(dup['Baris Data Cleansing'], '2, 3')
        with tempfile.TemporaryDirectory() as folder:
            bundle = Path(folder) / 'bundle.json'
            bundle.write_text(json.dumps(dict(output_columns=OUTPUT_COLUMNS, data_rows=rows,
                review_rows=reviews, provenance_rows=origins, balance_rows=ledger,
                summary_rows=[dict(Metrik='Record', Nilai=len(rows), Keterangan='Test')], assumptions=[])), encoding='utf-8')
            output = Path(folder) / 'output.xlsx'
            build_workbook(bundle, output)
            w = load_workbook(output)
            self.assertEqual(w['Data Cleansing']['B2'].fill.fgColor.rgb, '00F4CCCC')
            self.assertEqual(w['Data Cleansing']['B3'].fill.fgColor.rgb, '00F4CCCC')
            self.assertEqual(w['Data Cleansing']['B4'].fill.fgColor.rgb, '00FFF2CC')
            self.assertIn('AuditBalanceTable', w['Audit'].tables)
            self.assertIn('AuditProvenanceTable', w['Audit'].tables)
            self.assertIsNotNone(w['Data Cleansing'].tables['DataCleansingTable'].autoFilter)
            detail_table = w['Audit'].tables['AuditDuplicatesTable']
            self.assertIn('Tindak Lanjut', [column.name for column in detail_table.tableColumns])
            hyperlinks = [
                cell.hyperlink
                for row in w['Audit'].iter_rows()
                for cell in row
                if cell.hyperlink is not None
            ]
            self.assertLessEqual(len(hyperlinks), 20)
            self.assertTrue(all(link.location.startswith("'Audit'!") for link in hyperlinks))
            self.assertTrue(all(link.target is None for link in hyperlinks))
            w.close()


if __name__ == "__main__":
    unittest.main()
