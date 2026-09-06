"""Memory-bounded Excel export with native, independently filterable audit tables."""
from collections import Counter, defaultdict
from pathlib import Path
import warnings

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.hyperlink import Hyperlink

from .bundle import read_bundle
from .workbook import DATA_WIDTHS, AUDIT_WIDTHS, _partition_review_rows, _safe_value


COLORS = {'red': ('F4CCCC', '9C0006'), 'yellow': ('FFF2CC', '9C6500'),
          'purple': ('E4DFEC', '3F3151'), 'orange': ('FCE5CD', '783F04')}
ISSUE_LABELS = {
    'NAME_MULTIPLE_SAP': 'Nama sama, SAP berbeda',
    'SOURCE_DUPLICATE_SAP': 'SAP berulang dalam satu sumber',
    'SOURCE_DUPLICATE_ID': 'ID Vendor berulang dalam satu sumber',
    'SOURCE_DUPLICATE_NAME': 'Nama berulang dalam satu sumber',
    'SOURCE_RECORD_MISSING_SAP': 'SAP kosong pada sumber',
    'SOURCE_RECORD_MISSING_NAME': 'Nama kosong pada sumber',
    'SOURCE_RECORD_MISSING_ID': 'ID Vendor kosong pada sumber',
    'SOURCE_RECORD_MISSING_NPWP': 'NPWP kosong pada sumber',
    'SOURCE_RECORD_INVALID_NPWP': 'Format NPWP tidak valid',
    'NAME_CONFLICT': 'Nama berbeda untuk SAP sama',
    'NPWP_CONFLICT': 'NPWP berbeda untuk SAP sama',
    'ID_VENDOR_CONFLICT': 'ID Vendor berbeda untuk SAP sama',
    'CIRCLE_WITHOUT_PO': 'Belum ada PO terhubung',
    'NAME_LINK_REQUIRES_CONFIRMATION': 'Pencocokan nama perlu konfirmasi',
    'CURRENCY_LABEL_OVERRIDE': 'Label currency berbeda dari konfirmasi IDR',
    'PO_REPEATED_ITEM_KEY': 'Nomor PO dan item berulang',
    'PO_VENDOR_NO_REGISTRY_MATCH': 'SAP PO belum terhubung ke HK Circle',
    'PO_CIRCLE_NO_OVERLAP': 'Klasifikasi PO berbeda dari Circle',
    'PO_CIRCLE_PARTIAL_SUPPORT': 'Circle mendukung sebagian klasifikasi PO',
    'PO_RULE_GAP_CIRCLE_PRESENT': 'Final belum terdeteksi; Circle tersedia',
    'PO_RULE_GAP_CIRCLE_EMPTY': 'Final belum terdeteksi; Circle kosong',
    'CIRCLE_UNMAPPED': 'Deklarasi Circle belum dapat dibandingkan',
    'HIERARCHY_AMBIGUOUS_ITEM': 'Kelompok pekerjaan PO masih ambigu',
    'ITEM_TEXT_TRUNCATED': 'Daftar pekerjaan melebihi kapasitas satu sel',
    'AMBIGUOUS_EXACT_NAME': 'Nama sama mengarah ke beberapa SAP PO',
    'AMBIGUOUS_CANONICAL_NAME': 'Nama ternormalisasi mengarah ke beberapa SAP PO',
    'ID_TO_MULTIPLE_SAP': 'ID Vendor mengarah ke beberapa SAP',
}


def recommended_action(issue):
    if issue == 'CIRCLE_WITHOUT_PO':
        return 'Tidak perlu menghapus record. Periksa periode PO atau identitas penghubung bila pekerjaan seharusnya tersedia.'
    if issue == 'CURRENCY_LABEL_OVERRIDE':
        return 'Pastikan label sumber konsisten dengan konfirmasi IDR. Tidak ada konversi kurs otomatis.'
    if 'DUPLICATE' in issue or 'CONFLICT' in issue or 'MULTIPLE_SAP' in issue or 'AMBIGUOUS' in issue and not issue.startswith('HIERARCHY'):
        return 'Bandingkan identitas dan baris sumber terkait. Konfirmasi SAP yang benar; jangan gabungkan saldo atau hapus record berdasarkan nama saja.'
    if 'MISSING' in issue or 'INVALID' in issue:
        return 'Lengkapi atau koreksi data pada sumber HK Circle, lalu jalankan kembali otomasi.'
    if issue == 'PO_REPEATED_ITEM_KEY':
        return 'Konfirmasi apakah setiap baris adalah pekerjaan berbeda atau duplikasi ekspor. Seluruh nilai masih dihitung.'
    if issue == 'NAME_LINK_REQUIRES_CONFIRMATION':
        return 'Konfirmasi identitas calon dengan SAP pada bukti PO sebelum menerima klasifikasi sementara.'
    return 'Periksa uraian pekerjaan dan bukti pada PO; konfirmasi klasifikasi atau identitas sebelum memperbarui sumber/rule.'


def export_streaming(bundle_path: Path, output_path: Path):
    bundle = read_bundle(bundle_path)
    wb = Workbook(write_only=True)
    wb.properties.title = 'Data Cleansing Vendor Otomatis'
    wb.properties.creator = 'SCM Vendor Cleansing'
    fonts = {key: Font(name='Aptos', size=9, color=fg, bold=key in {'red', 'purple', 'orange'}) for key, (_, fg) in COLORS.items()}
    fills = {key: PatternFill('solid', fgColor=bg) for key, (bg, _) in COLORS.items()}
    normal_font = Font(name='Aptos', size=9, color='202020')
    header_font = Font(name='Aptos', size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='17365D')
    section_font = Font(name='Aptos', size=11, bold=True, color='17365D')
    section_fill = PatternFill('solid', fgColor='D9EAF7')
    alignment = Alignment(vertical='top', wrap_text=True)

    def cell(sheet, value, style='', field=''):
        value = _safe_value(value)
        c = WriteOnlyCell(sheet, value=value)
        c.font = normal_font
        c.alignment = alignment
        if isinstance(value, str):
            c.data_type = 's'  # Source values are data, never executable formulas.
        if style in COLORS:
            c.fill, c.font = fills[style], fonts[style]
        elif style == 'header':
            c.fill, c.font = header_fill, header_font
        elif style == 'section':
            c.fill, c.font = section_fill, section_font
        if field in {'ID Vendor', 'NO SAP', 'NPWP', 'SAP Sumber', 'SAP Bukti PO'}:
            c.number_format = '@'
        elif field == 'Saldo Hutang' or '(IDR)' in field:
            c.number_format = '#,##0.00'
        elif isinstance(value, (float, int)):
            c.number_format = '#,##0'
        # Detailed audit references stay as plain row numbers. Turning every
        # reference into a hyperlink can exceed Excel's per-worksheet limit
        # (65,530) and makes Excel remove all hyperlinks from the Audit sheet.
        return c

    def table(sheet, name, columns, header, count):
        if not count:
            return
        t = Table(displayName=name, ref=f'A{header}:{get_column_letter(len(columns))}{header + count}',
                  tableColumns=[TableColumn(id=i, name=column) for i, column in enumerate(columns, 1)])
        t.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
        t.autoFilter = AutoFilter(ref=t.ref)
        # openpyxl warns unconditionally for write-only tables; explicit columns
        # above satisfy the requirement and are verified on the saved XLSX.
        with warnings.catch_warnings():
            warnings.filterwarnings('ignore', message='In write-only mode you must add table columns manually')
            sheet.add_table(t)

    data = wb.create_sheet('Data Cleansing')
    data.sheet_view.showGridLines = False
    data.sheet_view.zoomScale = 80
    data.freeze_panes = 'D2'
    for i, width in enumerate(DATA_WIDTHS, 1):
        data.column_dimensions[get_column_letter(i)].width = width
    highlights = defaultdict(dict)
    sap_rows = defaultdict(list)
    for number, row in enumerate(bundle['data_rows'], 2):
        if row['NO SAP']:
            sap_rows[str(row['NO SAP'])].append(number)
    fields = {'ID_VENDOR_CONFLICT': 1, 'NPWP_CONFLICT': 4, 'NAME_CONFLICT': 3,
        'NAME_MULTIPLE_SAP': 3, 'SOURCE_RECORD_MISSING_SAP': 2, 'SOURCE_RECORD_MISSING_NAME': 3,
        'SOURCE_RECORD_MISSING_ID': 1, 'MISSING_ID_VENDOR': 1, 'SOURCE_RECORD_MISSING_NPWP': 4,
        'MISSING_NPWP': 4, 'SOURCE_RECORD_INVALID_NPWP': 4, 'PO_CIRCLE_NO_OVERLAP': 19,
        'CIRCLE_UNMAPPED': 19, 'NAME_LINK_REQUIRES_CONFIRMATION': 20, 'CURRENCY_LABEL_OVERRIDE': 25}
    source_cols = {'DRT': 7, 'DM': 7, 'DRT_LAMA': 8, 'DM_LAMA': 8, 'DCR': 10, 'DCM': 11, 'DBCR': 12}
    for review in bundle.get('review_rows', []):
        numbers = [int(x.strip()) for x in str(review.get('Baris Data Cleansing', '')).split(',') if x.strip().isdigit()]
        if 'Baris Data Cleansing' not in review:
            numbers = sap_rows.get(str(review.get('NO SAP', '')), [])
        issue = review.get('Issue', '')
        red = (review.get('Severity') == 'HIGH' and 'MISSING' not in issue) or 'DUPLICATE' in issue
        for number in numbers:
            if not 2 <= number <= len(bundle['data_rows']) + 1:
                raise ValueError(f'Baris audit di luar hasil: {number}')
            colors = highlights[number]
            if red:
                colors[2] = 'red'
            field = fields.get(issue)
            if field and colors.get(field) != 'red':
                colors[field] = 'red' if red else 'yellow'
            if 'DUPLICATE' in issue and review.get('Source') in source_cols:
                colors[source_cols[review['Source']]] = 'red'
    columns = bundle['output_columns']
    data.row_dimensions[1].height = 72
    data.append([cell(data, col, 'header') for col in columns])
    for number, row in enumerate(bundle['data_rows'], 2):
        styles = {i: 'yellow' for i in (1, 2, 3, 4) if not row[columns[i - 1]]}
        for i in (20, 22, 23, 24):
            if not row[columns[i - 1]]:
                styles[i] = 'purple'
        if row.get('Inject') == '✓':
            styles[9] = 'orange'
        styles.update(highlights.get(number, {}))
        data.row_dimensions[number].height = 42
        data.append([cell(data, row.get(col, ''), styles.get(i, ''), col) for i, col in enumerate(columns, 1)])
    table(data, 'DataCleansingTable', columns, 1, len(bundle['data_rows']))

    audit = wb.create_sheet('Audit')
    audit.sheet_view.showGridLines = False
    audit.sheet_view.zoomScale = 85
    audit.freeze_panes = 'C4'
    for i, width in enumerate(AUDIT_WIDTHS, 1):
        audit.column_dimensions[get_column_letter(i)].width = width
    reviews = bundle.get('review_rows', [])
    rank = {'HIGH': 0, 'MEDIUM': 1, 'LOW': 2}
    counts = Counter((r.get('Severity', ''), r.get('Issue', '')) for r in reviews)
    issue_summary = [{'Severity': severity, 'Jenis Temuan': ISSUE_LABELS.get(issue, issue), 'Jumlah Temuan': count,
        'Arti Jumlah': 'Jumlah kejadian atau kelompok temuan, bukan jumlah vendor unik.', 'Kode Temuan': issue}
        for (severity, issue), count in sorted(counts.items(), key=lambda x: (rank.get(x[0][0], 3), -x[1], x[0][1]))]
    sections = [('Ringkasan Jenis Temuan', 'AuditIssueSummaryTable',
                 ['Severity', 'Jenis Temuan', 'Jumlah Temuan', 'Arti Jumlah', 'Kode Temuan'], issue_summary)]
    partitions = _partition_review_rows(reviews)
    review_columns = ['Severity', 'Jenis Temuan', 'Source', 'Source Row', 'ID Vendor', 'NO SAP', 'Nama Rekanan', 'Match Method', 'Detail', 'Baris Data Cleansing', 'Issue', 'Tindak Lanjut']
    for title, key, name in [
        ('Duplikasi dan Konflik Identitas', 'duplicates', 'AuditDuplicatesTable'),
        ('Kelengkapan dan Format Data', 'completeness', 'AuditCompletenessTable'),
        ('Pencocokan HK Circle terhadap PO', 'matching', 'AuditMatchingTable'),
        ('Perbedaan dan Ketidakpastian Klasifikasi', 'classification', 'AuditClassificationReviewTable'),
        ('Temuan Lain', 'other', 'AuditOtherFindingsTable')]:
        detail_rows = sorted(partitions[key], key=lambda r: (rank.get(r.get('Severity'), 3), r.get('Issue', ''), r.get('Source', ''), str(r.get('Source Row', ''))))
        if detail_rows:
            sections.append((title, name, review_columns, detail_rows))
    for title, key, name, cols in [
        ('Akumulasi Nilai PO per SAP dalam IDR', 'balance_rows', 'AuditBalanceTable',
         ['NO SAP', 'Nama Rekanan', 'Baris Data Cleansing', 'Jumlah Item PO', 'Nilai PO HK (IDR)', 'Nilai PO JO (IDR)', 'Saldo Hutang (IDR)', 'Currency Sumber', 'Dasar']),
        ('Penelusuran Record dan Pemilik Saldo', 'provenance_rows', 'AuditProvenanceTable',
         ['Baris Data Cleansing', 'Source', 'Source File', 'Source Row', 'ID Vendor', 'NO SAP', 'Nama Rekanan', 'Match Method', 'Status Saldo', 'Baris Pemilik Saldo', 'SAP Sumber', 'SAP Bukti PO']),
        ('Bukti Klasifikasi Final dari PO', 'evidence_rows', 'AuditClassificationEvidenceTable',
         ['NO SAP', 'Nama Vendor PO', 'Rank', 'Klasifikasi', 'Jumlah PO Berbeda', 'Jumlah Item PO', 'Rule ID', 'Confidence Rule', 'Dukungan Circle', 'Sumber Final', 'Contoh Deskripsi']),
        ('Bukti Kelompok Level 1–3 dari Item PO', 'hierarchy_evidence_rows', 'AuditHierarchyEvidenceTable',
         ['NO SAP', 'Rank', 'Level 1', 'Kode Level 2', 'Level 2', 'Level 3', 'Jumlah PO Berbeda', 'Jumlah Item PO', 'Baris Sumber PO', 'Bukti Istilah', 'Contoh Deskripsi', 'Boundary Diterapkan']),
        ('Item PO Belum Memiliki Kelompok Level 1–3', 'hierarchy_unresolved_rows', 'AuditHierarchyUnresolvedTable',
         ['Company', 'NO SAP', 'Nama Vendor', 'Deskripsi Belum Memiliki Level', 'Alasan', 'Detail', 'Jumlah Item', 'Contoh PO', 'Contoh Item PO', 'Baris Sumber PO', 'Contoh Project', 'Tindakan']),
        ('Item PO Belum Terklasifikasi', 'unresolved_rows', 'AuditUnresolvedPOTable',
         ['Company', 'NO SAP', 'Nama Vendor', 'Deskripsi Belum Terklasifikasi', 'Jumlah Item', 'Contoh PO', 'Contoh Item PO', 'Contoh Project', 'Tindakan'])]:
        detail_rows = bundle.get(key, [])
        if detail_rows:
            sections.append((title, name, cols, detail_rows))

    summary = bundle.get('summary_rows', [])
    assumptions = bundle.get('assumptions', [])
    start = max(len(summary) + 6, len(assumptions) + len(sections) + 17)
    starts = []
    cursor = start
    for title, name, cols, detail_rows in sections:
        starts.append(cursor)
        cursor += len(detail_rows) + 4
    if cursor > 1048576:
        raise ValueError('Audit melebihi kapasitas baris Excel; persempit periode sumber.')
    top = defaultdict(dict)
    top[1][1] = cell(audit, 'Laporan Audit HK Circle', 'section')
    for i, value in enumerate(['Metrik', 'Nilai', 'Keterangan'], 1):
        top[3][i] = cell(audit, value, 'header')
    for number, record in enumerate(summary, 4):
        for i, key in enumerate(['Metrik', 'Nilai', 'Keterangan'], 1):
            top[number][i] = cell(audit, record[key])
    severity = Counter(r.get('Severity') for r in reviews)
    top[3][5], top[3][6] = cell(audit, 'Severity', 'header'), cell(audit, 'Jumlah', 'header')
    for number, label in enumerate(['HIGH', 'MEDIUM', 'LOW', 'TOTAL'], 4):
        top[number][5] = cell(audit, label)
        top[number][6] = cell(audit, len(reviews) if label == 'TOTAL' else severity[label])
    top[3][8], top[3][9] = cell(audit, 'Warna', 'header'), cell(audit, 'Arti Highlight', 'header')
    for number, (color, label, meaning) in enumerate([
        ('red', 'Merah', 'Duplikasi atau konflik HIGH. Record tidak dihapus.'),
        ('yellow', 'Kuning', 'Data kosong/tidak valid, atau identitas/currency perlu konfirmasi.'),
        ('purple', 'Ungu', 'Klasifikasi Final atau Level 1–3 belum memiliki bukti yang cukup.'),
        ('orange', 'Oranye', 'Inject: mempunyai PO, belum terdaftar pada master aktif/lama. Bukan tingkat anomali.')], 4):
        top[number][8], top[number][9] = cell(audit, label, color), cell(audit, meaning)
    top[9][5] = cell(audit, 'Aturan dan cara membaca laporan', 'section')
    for number, text in enumerate(assumptions, 10):
        # Use a dedicated wide column; no overflow-dependent merged cells.
        top[number][9] = cell(audit, text)
    nav = len(assumptions) + 12
    top[nav][9] = cell(audit, 'Navigasi — klik untuk menuju tabel', 'section')
    for number, ((title, _, _, detail_rows), position) in enumerate(zip(sections, starts), nav + 1):
        c = cell(audit, f'{title} ({len(detail_rows):,} baris)')
        # Use an internal location. A relationship with TargetMode=External
        # for a local sheet address is tolerated by some readers but can be
        # repaired away by stricter Excel versions.
        c.hyperlink = Hyperlink(ref='', location=f"'Audit'!A{position}")
        c.font = Font(name='Aptos', size=9, color='0563C1', underline='single')
        top[number][9] = c
    for number in range(1, start):
        audit.row_dimensions[number].height = 72 if 10 <= number < 10 + len(assumptions) else 48
        audit.append([top[number].get(i, cell(audit, '')) for i in range(1, 10)])
    table(audit, 'AuditSummaryTable', ['Metrik', 'Nilai', 'Keterangan'], 3, len(summary))
    # Severity lives at E:F, so add its explicit native table separately.
    t = Table(displayName='AuditSeverityTable', ref='E3:F7', tableColumns=[TableColumn(id=1, name='Severity'), TableColumn(id=2, name='Jumlah')])
    t.tableStyleInfo = TableStyleInfo(name='TableStyleMedium2', showRowStripes=True)
    t.autoFilter = AutoFilter(ref=t.ref)
    with warnings.catch_warnings():
        warnings.filterwarnings('ignore', message='In write-only mode you must add table columns manually')
        audit.add_table(t)
    cursor = start
    for (title, name, cols, detail_rows), position in zip(sections, starts):
        audit.row_dimensions[cursor].height = 30
        title_cell = cell(audit, f'{title} ({len(detail_rows):,} baris)', 'section')
        title_cell.alignment = Alignment(vertical='center', wrap_text=False)
        audit.append([title_cell])
        cursor += 1
        audit.row_dimensions[cursor].height = 48
        audit.append([cell(audit, col, 'header') for col in cols])
        table(audit, name, cols, cursor, len(detail_rows))
        cursor += 1
        for record in detail_rows:
            values = []
            for col in cols:
                value = ISSUE_LABELS.get(record.get('Issue'), record.get('Issue', '')) if col == 'Jenis Temuan' and 'Issue' in record else record.get(col, '')
                if col == 'Tindak Lanjut':
                    value = recommended_action(record.get('Issue', ''))
                style = {'HIGH': 'red', 'MEDIUM': 'yellow'}.get(value, '') if col == 'Severity' else ''
                values.append(cell(audit, value, style, col))
            audit.row_dimensions[cursor].height = 54
            audit.append(values)
            cursor += 1
        audit.append([])
        audit.append([])
        cursor += 2
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
