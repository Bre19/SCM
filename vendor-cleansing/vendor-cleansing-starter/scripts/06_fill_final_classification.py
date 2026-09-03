from __future__ import annotations

import shutil
import sys
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]

if str(ROOT) not in sys.path:
    sys.path.insert(
        0,
        str(ROOT),
    )


from src.classifier import (
    classify_text,
    load_classification_rules,
    load_enabled_vocabulary,
    normalize_id,
    normalize_text,
)

from src.config import (
    APPROVED_VOCABULARY_FILE,
    AUDIT_REVIEW_FILE,
    CLASSIFICATION_RULES_FILE,
    CLEANSING_SHEET,
    EXPECTED_VENDOR_ROWS,
    FINAL_OUTPUT_DIR,
    FINAL_OUTPUT_FILE,
    PO_END_DATE,
    PO_FILE,
    PO_HEADER,
    PO_SHEET,
    PO_START_DATE,
    PREVIEW_OUTPUT_FILE,
    REQUIRE_ALL_CLASSIFIED_FOR_FINAL,
    REQUIRED_PO_COLUMNS,
    WRITE_PREVIEW_ONLY,
)


TARGET_FINAL_COLUMN = (
    "Klasifikasi Final"
)

VENDOR_ID_COLUMN = (
    "NO SAP"
)

VENDOR_NAME_COLUMN = (
    "Nama Rekanan"
)

PO_ITEM_COLUMN = (
    "Item Pekerjaan Berdasarkan PO"
)

LEVEL_1_COLUMN = (
    "Kelompok Klasifikasi\n"
    "Level 1 (Klasifikasi PO)"
)

LEVEL_2_COLUMN = (
    "Kelompok Klasifikasi\n"
    "Level 2 (Klasifikasi PO)"
)

LEVEL_3_COLUMN = (
    "Kelompok Klasifikasi\n"
    "Level 3 (Klasifikasi PO)"
)


def require_columns(
    df: pd.DataFrame,
    required: list[str],
    source: str,
) -> None:
    missing = [
        column
        for column in required
        if column not in df.columns
    ]

    if missing:
        raise ValueError(
            f"{source}: missing columns: "
            f"{missing}"
        )


def build_po_classification(
    po: pd.DataFrame,
    rules,
) -> pd.DataFrame:
    evidence: list[
        dict[str, object]
    ] = []

    for _, row in po.iterrows():
        vendor_id = normalize_id(
            row["Vendor"]
        )

        po_number = normalize_id(
            row["PO"]
        )

        description = (
            row["Deskripsi"]
        )

        matches = classify_text(
            description,
            rules,
        )

        for (
            classification,
            rule,
        ) in matches:

            evidence.append(
                {
                    "vendor_id":
                        vendor_id,
                    "po":
                        po_number,
                    "item_po":
                        normalize_id(
                            row["Item.PO"]
                        ),
                    "description":
                        str(
                            description
                        ).strip(),
                    "classification":
                        classification,
                    "rule_priority":
                        rule.priority,
                    "rule_pattern":
                        rule.pattern_text,
                }
            )

    if not evidence:
        return pd.DataFrame(
            columns=[
                "vendor_id",
                "po",
                "item_po",
                "description",
                "classification",
                "rule_priority",
                "rule_pattern",
            ]
        )

    return pd.DataFrame(
        evidence
    )


def aggregate_po_classification(
    evidence: pd.DataFrame,
) -> pd.DataFrame:
    if evidence.empty:
        return pd.DataFrame(
            columns=[
                "vendor_id",
                "classification",
                "po_count",
                "item_count",
                "examples",
            ]
        )

    def examples(
        values: pd.Series,
    ) -> str:
        unique = list(
            dict.fromkeys(
                str(value).strip()
                for value in values
                if str(value).strip()
            )
        )

        return " | ".join(
            unique[:5]
        )

    result = (
        evidence
        .groupby(
            [
                "vendor_id",
                "classification",
            ],
            as_index=False,
        )
        .agg(
            po_count=(
                "po",
                "nunique",
            ),
            item_count=(
                "item_po",
                "size",
            ),
            examples=(
                "description",
                examples,
            ),
        )
    )

    result = result.sort_values(
        [
            "vendor_id",
            "po_count",
            "item_count",
            "classification",
        ],
        ascending=[
            True,
            False,
            False,
            True,
        ],
    )

    return result


def create_vendor_rankings(
    aggregated: pd.DataFrame,
) -> dict[
    str,
    list[
        dict[str, object]
    ],
]:
    result: dict[
        str,
        list[
            dict[str, object]
        ],
    ] = defaultdict(
        list
    )

    for _, row in (
        aggregated.iterrows()
    ):
        result[
            row["vendor_id"]
        ].append(
            {
                "classification":
                    row[
                        "classification"
                    ],
                "po_count":
                    int(
                        row["po_count"]
                    ),
                "item_count":
                    int(
                        row["item_count"]
                    ),
                "examples":
                    row["examples"],
                "source":
                    "DATA_PO",
            }
        )

    return dict(
        result
    )


def build_unique_name_mapping(
    po: pd.DataFrame,
) -> tuple[
    dict[str, str],
    set[str],
]:
    name_to_ids: dict[
        str,
        set[str],
    ] = defaultdict(
        set
    )

    for _, row in po.iterrows():
        name = normalize_text(
            row["Nama Vendor"]
        )

        vendor_id = normalize_id(
            row["Vendor"]
        )

        if (
            name
            and vendor_id
        ):
            name_to_ids[
                name
            ].add(
                vendor_id
            )

    unique: dict[
        str,
        str,
    ] = {}

    ambiguous: set[
        str
    ] = set()

    for (
        name,
        vendor_ids,
    ) in name_to_ids.items():

        if len(
            vendor_ids
        ) == 1:
            unique[
                name
            ] = next(
                iter(
                    vendor_ids
                )
            )

        elif len(
            vendor_ids
        ) > 1:
            ambiguous.add(
                name
            )

    return (
        unique,
        ambiguous,
    )


def classify_fallback_from_cleansing(
    row: pd.Series,
    rules,
) -> list[str]:
    combined = " | ".join(
        [
            str(
                row.get(
                    PO_ITEM_COLUMN,
                    "",
                )
            ),
            str(
                row.get(
                    LEVEL_1_COLUMN,
                    "",
                )
            ),
            str(
                row.get(
                    LEVEL_2_COLUMN,
                    "",
                )
            ),
            str(
                row.get(
                    LEVEL_3_COLUMN,
                    "",
                )
            ),
        ]
    )

    matches = classify_text(
        combined,
        rules,
    )

    return [
        classification
        for (
            classification,
            _,
        ) in matches
    ]


def find_header_column(
    worksheet,
    header_name: str,
) -> int:
    for cell in worksheet[1]:
        if (
            str(
                cell.value
            ).strip()
            == header_name
        ):
            return int(
                cell.column
            )

    raise ValueError(
        f"Header not found: "
        f"{header_name}"
    )


def main() -> None:
    FINAL_OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    # =====================================================
    # VOCABULARY + RULES
    # =====================================================

    enabled_vocabulary = (
        load_enabled_vocabulary(
            APPROVED_VOCABULARY_FILE
        )
    )

    rules = (
        load_classification_rules(
            CLASSIFICATION_RULES_FILE,
            enabled_vocabulary,
        )
    )

    # =====================================================
    # LOAD AUDITED DATA CLEANSING
    # =====================================================

    cleansing = pd.read_excel(
        AUDIT_REVIEW_FILE,
        sheet_name=CLEANSING_SHEET,
        dtype=str,
        keep_default_na=False,
    )

    if (
        len(cleansing)
        != EXPECTED_VENDOR_ROWS
    ):
        raise RuntimeError(
            "Data Cleansing row guard failed: "
            f"expected "
            f"{EXPECTED_VENDOR_ROWS:,}, "
            f"got {len(cleansing):,}"
        )

    # =====================================================
    # LOAD DATA PO
    # =====================================================

    po = pd.read_excel(
        PO_FILE,
        sheet_name=PO_SHEET,
        header=PO_HEADER,
        dtype=str,
        keep_default_na=False,
    )

    require_columns(
        po,
        REQUIRED_PO_COLUMNS,
        "Data PO",
    )

    po["Doc.Date"] = pd.to_datetime(
        po["Doc.Date"],
        errors="coerce",
        dayfirst=True,
    )

    start_date = pd.Timestamp(
        PO_START_DATE
    )

    end_date = pd.Timestamp(
        PO_END_DATE
    )

    po = po.loc[
        po["Doc.Date"].between(
            start_date,
            end_date,
            inclusive="both",
        )
    ].copy()

    # =====================================================
    # CLASSIFY ALL PO ITEMS
    # =====================================================

    evidence = (
        build_po_classification(
            po,
            rules,
        )
    )

    aggregated = (
        aggregate_po_classification(
            evidence
        )
    )

    vendor_rankings = (
        create_vendor_rankings(
            aggregated
        )
    )

    (
        unique_name_mapping,
        ambiguous_names,
    ) = (
        build_unique_name_mapping(
            po
        )
    )

    # =====================================================
    # BUILD PROPOSALS
    # =====================================================

    proposals: list[
        dict[str, object]
    ] = []

    detailed_evidence: list[
        dict[str, object]
    ] = []

    unresolved: list[
        dict[str, object]
    ] = []

    existing_count = 0
    target_count = 0

    for dataframe_index, row in (
        cleansing.iterrows()
    ):
        excel_row = (
            int(
                dataframe_index
            )
            + 2
        )

        existing_final = str(
            row[
                TARGET_FINAL_COLUMN
            ]
        ).strip()

        if existing_final:
            existing_count += 1

            proposals.append(
                {
                    "excel_row":
                        excel_row,
                    "no_sap":
                        row[
                            VENDOR_ID_COLUMN
                        ],
                    "vendor":
                        row[
                            VENDOR_NAME_COLUMN
                        ],
                    "status":
                        "EXISTING_PRESERVED",
                    "proposed_final":
                        existing_final,
                    "match_vendor_id":
                        normalize_id(
                            row[
                                VENDOR_ID_COLUMN
                            ]
                        ),
                }
            )

            continue

        target_count += 1

        source_vendor_id = normalize_id(
            row[
                VENDOR_ID_COLUMN
            ]
        )

        vendor_name_key = normalize_text(
            row[
                VENDOR_NAME_COLUMN
            ]
        )

        matched_vendor_id = (
            source_vendor_id
        )

        match_method = (
            "NO_SAP"
            if source_vendor_id
            else ""
        )

        if not matched_vendor_id:
            matched_vendor_id = (
                unique_name_mapping.get(
                    vendor_name_key,
                    "",
                )
            )

            if matched_vendor_id:
                match_method = (
                    "UNIQUE_VENDOR_NAME"
                )

        ranking = list(
            vendor_rankings.get(
                matched_vendor_id,
                [],
            )
        )

        labels = [
            item[
                "classification"
            ]
            for item in ranking
        ]

        # =============================================
        # FALLBACK:
        # gunakan U/V/W/X jika raw PO rule belum
        # menangkap klasifikasi tertentu.
        # =============================================

        fallback_labels = (
            classify_fallback_from_cleansing(
                row,
                rules,
            )
        )

        for label in fallback_labels:
            if label not in labels:
                labels.append(
                    label
                )

                ranking.append(
                    {
                        "classification":
                            label,
                        "po_count":
                            0,
                        "item_count":
                            0,
                        "examples":
                            "Fallback dari U/V/W/X",
                        "source":
                            (
                                "DATA_CLEANSING_"
                                "FALLBACK"
                            ),
                    }
                )

        proposed_final = (
            ", ".join(
                labels
            )
        )

        if proposed_final:
            status = (
                "PROPOSED"
            )

        else:
            status = (
                "UNRESOLVED"
            )

            reason = (
                "No classification rule matched"
            )

            if (
                not source_vendor_id
                and vendor_name_key
                in ambiguous_names
            ):
                reason = (
                    "NO SAP kosong dan "
                    "Nama Rekanan match ke "
                    "lebih dari satu Vendor SAP"
                )

            unresolved.append(
                {
                    "excel_row":
                        excel_row,
                    "no_sap":
                        row[
                            VENDOR_ID_COLUMN
                        ],
                    "vendor":
                        row[
                            VENDOR_NAME_COLUMN
                        ],
                    "reason":
                        reason,
                    "item_pekerjaan":
                        row.get(
                            PO_ITEM_COLUMN,
                            "",
                        ),
                    "level_1":
                        row.get(
                            LEVEL_1_COLUMN,
                            "",
                        ),
                    "level_2":
                        row.get(
                            LEVEL_2_COLUMN,
                            "",
                        ),
                    "level_3":
                        row.get(
                            LEVEL_3_COLUMN,
                            "",
                        ),
                }
            )

        proposals.append(
            {
                "excel_row":
                    excel_row,
                "no_sap":
                    row[
                        VENDOR_ID_COLUMN
                    ],
                "vendor":
                    row[
                        VENDOR_NAME_COLUMN
                    ],
                "status":
                    status,
                "proposed_final":
                    proposed_final,
                "match_vendor_id":
                    matched_vendor_id,
                "match_method":
                    match_method,
            }
        )

        for rank_number, item in enumerate(
            ranking,
            start=1,
        ):
            detailed_evidence.append(
                {
                    "excel_row":
                        excel_row,
                    "no_sap":
                        row[
                            VENDOR_ID_COLUMN
                        ],
                    "vendor":
                        row[
                            VENDOR_NAME_COLUMN
                        ],
                    "rank":
                        rank_number,
                    "classification":
                        item[
                            "classification"
                        ],
                    "distinct_po_count":
                        item[
                            "po_count"
                        ],
                    "item_count":
                        item[
                            "item_count"
                        ],
                    "source":
                        item[
                            "source"
                        ],
                    "examples":
                        item[
                            "examples"
                        ],
                }
            )

    proposal_df = pd.DataFrame(
        proposals
    )

    evidence_df = pd.DataFrame(
        detailed_evidence
    )

    unresolved_df = pd.DataFrame(
        unresolved
    )

    # =====================================================
    # SAVE CSV
    # =====================================================

    proposal_df.to_csv(
        FINAL_OUTPUT_DIR
        / "classification_proposals.csv",
        index=False,
        encoding="utf-8-sig",
    )

    evidence_df.to_csv(
        FINAL_OUTPUT_DIR
        / "classification_evidence.csv",
        index=False,
        encoding="utf-8-sig",
    )

    unresolved_df.to_csv(
        FINAL_OUTPUT_DIR
        / "classification_unresolved.csv",
        index=False,
        encoding="utf-8-sig",
    )

    # =====================================================
    # PREVIEW / FINAL WORKBOOK
    # =====================================================

    if WRITE_PREVIEW_ONLY:
        output_file = (
            PREVIEW_OUTPUT_FILE
        )
    else:
        output_file = (
            FINAL_OUTPUT_FILE
        )

    shutil.copy2(
        AUDIT_REVIEW_FILE,
        output_file,
    )

    workbook = load_workbook(
        output_file
    )

    worksheet = workbook[
        CLEANSING_SHEET
    ]

    final_column = (
        find_header_column(
            worksheet,
            TARGET_FINAL_COLUMN,
        )
    )

    # =====================================================
    # WRITE ONLY BLANK FINAL CLASSIFICATION
    # =====================================================

    proposal_lookup = (
        proposal_df
        .set_index(
            "excel_row"
        )
    )

    review_fill = PatternFill(
        fill_type="solid",
        fgColor="D9D2E9",
    )

    for excel_row, proposal in (
        proposal_lookup.iterrows()
    ):
        cell = worksheet.cell(
            row=int(
                excel_row
            ),
            column=final_column,
        )

        # Existing value MUST NOT be overwritten.
        if str(
            cell.value or ""
        ).strip():
            continue

        proposed_final = str(
            proposal[
                "proposed_final"
            ]
        ).strip()

        if proposed_final:
            cell.value = (
                proposed_final
            )

        else:
            # Unresolved classification:
            # jangan mengarang label.
            # Tandai T agar mudah direview.
            cell.fill = review_fill
            cell.font = Font(
                bold=True,
            )

    # =====================================================
    # CLASSIFICATION EVIDENCE SHEET
    # =====================================================

    if (
        "Classification Evidence"
        in workbook.sheetnames
    ):
        del workbook[
            "Classification Evidence"
        ]

    evidence_ws = (
        workbook.create_sheet(
            "Classification Evidence"
        )
    )

    evidence_headers = [
        "Excel Row",
        "NO SAP",
        "Vendor",
        "Rank",
        "Classification",
        "Distinct PO Count",
        "Item Count",
        "Source",
        "Examples",
    ]

    evidence_ws.append(
        evidence_headers
    )

    for cell in evidence_ws[1]:
        cell.font = Font(
            bold=True
        )

    if not evidence_df.empty:
        for _, row in (
            evidence_df.iterrows()
        ):
            evidence_ws.append(
                [
                    row[
                        "excel_row"
                    ],
                    row[
                        "no_sap"
                    ],
                    row[
                        "vendor"
                    ],
                    row[
                        "rank"
                    ],
                    row[
                        "classification"
                    ],
                    row[
                        "distinct_po_count"
                    ],
                    row[
                        "item_count"
                    ],
                    row[
                        "source"
                    ],
                    row[
                        "examples"
                    ],
                ]
            )

    evidence_ws.freeze_panes = (
        "A2"
    )

    evidence_ws.auto_filter.ref = (
        evidence_ws.dimensions
    )

    # =====================================================
    # CLASSIFICATION REVIEW SHEET
    # =====================================================

    if (
        "Classification Review"
        in workbook.sheetnames
    ):
        del workbook[
            "Classification Review"
        ]

    review_ws = workbook.create_sheet(
        "Classification Review"
    )

    review_headers = [
        "Excel Row",
        "NO SAP",
        "Vendor",
        "Reason",
        "Item Pekerjaan",
        "Level 1",
        "Level 2",
        "Level 3",
    ]

    review_ws.append(
        review_headers
    )

    for cell in review_ws[1]:
        cell.font = Font(
            bold=True
        )

    if not unresolved_df.empty:
        for _, row in (
            unresolved_df.iterrows()
        ):
            review_ws.append(
                [
                    row[
                        "excel_row"
                    ],
                    row[
                        "no_sap"
                    ],
                    row[
                        "vendor"
                    ],
                    row[
                        "reason"
                    ],
                    row[
                        "item_pekerjaan"
                    ],
                    row[
                        "level_1"
                    ],
                    row[
                        "level_2"
                    ],
                    row[
                        "level_3"
                    ],
                ]
            )

    review_ws.freeze_panes = (
        "A2"
    )

    review_ws.auto_filter.ref = (
        review_ws.dimensions
    )

    workbook.save(
        output_file
    )

    # =====================================================
    # FINAL GUARDS
    # =====================================================

    verify = pd.read_excel(
        output_file,
        sheet_name=CLEANSING_SHEET,
        dtype=str,
        keep_default_na=False,
    )

    if (
        len(verify)
        != EXPECTED_VENDOR_ROWS
    ):
        raise RuntimeError(
            "OUTPUT ROW GUARD FAILED: "
            f"expected "
            f"{EXPECTED_VENDOR_ROWS:,}, "
            f"got {len(verify):,}"
        )

    remaining_blank = int(
        verify[
            TARGET_FINAL_COLUMN
        ]
        .astype(str)
        .str.strip()
        .eq("")
        .sum()
    )

    newly_classified = (
        target_count
        - len(
            unresolved_df
        )
    )

    print()
    print("=" * 72)
    print(
        "STAGE 6 - FINAL CLASSIFICATION"
    )
    print("=" * 72)

    print(
        f"Vendor rows             : "
        f"{len(cleansing):,}"
    )

    print(
        f"Existing final preserved: "
        f"{existing_count:,}"
    )

    print(
        f"Blank final target      : "
        f"{target_count:,}"
    )

    print(
        f"Newly classified        : "
        f"{newly_classified:,}"
    )

    print(
        f"Unresolved              : "
        f"{len(unresolved_df):,}"
    )

    print(
        f"Remaining blank T       : "
        f"{remaining_blank:,}"
    )

    print(
        f"PO evidence rows        : "
        f"{len(evidence):,}"
    )

    print(
        f"Output                   : "
        f"{output_file}"
    )

    print("=" * 72)

    if (
        not WRITE_PREVIEW_ONLY
        and REQUIRE_ALL_CLASSIFIED_FOR_FINAL
        and remaining_blank > 0
    ):
        raise RuntimeError(
            "FINAL OUTPUT REJECTED: "
            f"{remaining_blank:,} "
            "Klasifikasi Final masih kosong. "
            "Perbaiki classification_rules.csv "
            "dan jalankan ulang."
        )


if __name__ == "__main__":
    main()