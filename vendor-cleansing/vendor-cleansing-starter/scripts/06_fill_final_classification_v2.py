from __future__ import annotations

import shutil
import sys
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = (
    Path(__file__)
    .resolve()
    .parents[1]
)

if str(ROOT) not in sys.path:
    sys.path.insert(
        0,
        str(ROOT),
    )


from src.classifier_v2 import (
    classify_context,
    classify_po_description,
    load_context_rules,
    load_po_rules,
    load_vocabulary,
    normalize_id,
    normalize_text,
)

from src.config import (
    AUDIT_REVIEW_FILE,
    CLEANSING_SHEET,
    CONTEXT_RULES_V2_FILE,
    EXPECTED_VENDOR_ROWS,
    NEW_VOCABULARY_FILL,
    PO_END_DATE,
    PO_FILE,
    PO_HEADER,
    PO_RULES_V2_FILE,
    PO_SHEET,
    PO_START_DATE,
    REQUIRED_PO_COLUMNS,
    UNRESOLVED_CLASSIFICATION_FILL,
    V2_FINAL_OUTPUT_FILE,
    V2_OUTPUT_DIR,
    V2_PREVIEW_OUTPUT_FILE,
    VOCABULARY_V2_FILE,
    WRITE_V2_PREVIEW_ONLY,
)


FINAL_COLUMN = (
    "Klasifikasi Final"
)

SAP_COLUMN = (
    "NO SAP"
)

VENDOR_COLUMN = (
    "Nama Rekanan"
)

CIRCLE_COLUMN = (
    "Klasifikasi Circle"
)

PO_ITEM_COLUMN = (
    "Item Pekerjaan Berdasarkan PO"
)

LEVEL_2_COLUMN = (
    "Kelompok Klasifikasi\n"
    "Level 2 (Klasifikasi PO)"
)

LEVEL_3_COLUMN = (
    "Kelompok Klasifikasi\n"
    "Level 3 (Klasifikasi PO)"
)


def require_columns(
    df: pd.DataFrame,
    required: list[str],
    source: str,
) -> None:
    missing = [
        column
        for column in required
        if column not in df.columns
    ]

    if missing:
        raise ValueError(
            f"{source}: "
            f"missing columns: {missing}"
        )


def find_header_column(
    worksheet,
    header_name: str,
) -> int:
    for cell in worksheet[1]:
        if (
            str(
                cell.value
            ).strip()
            == header_name
        ):
            return int(
                cell.column
            )

    raise ValueError(
        f"Header not found: "
        f"{header_name}"
    )


def build_po_evidence(
    po: pd.DataFrame,
    po_rules,
) -> pd.DataFrame:
    rows: list[
        dict[str, object]
    ] = []

    for _, row in (
        po.iterrows()
    ):
        vendor_id = normalize_id(
            row["Vendor"]
        )

        po_number = normalize_id(
            row["PO"]
        )

        item_po = normalize_id(
            row["Item.PO"]
        )

        description = str(
            row["Deskripsi"]
        ).strip()

        matches = (
            classify_po_description(
                description,
                po_rules,
            )
        )

        for rule in matches:
            rows.append(
                {
                    "vendor_id":
                        vendor_id,
                    "po":
                        po_number,
                    "item_po":
                        item_po,
                    "description":
                        description,
                    "classification":
                        rule.classification,
                    "rule_priority":
                        rule.priority,
                    "rule_pattern":
                        rule.pattern_text,
                }
            )

    return pd.DataFrame(
        rows
    )


def aggregate_po_evidence(
    evidence: pd.DataFrame,
) -> pd.DataFrame:
    if evidence.empty:
        return pd.DataFrame(
            columns=[
                "vendor_id",
                "classification",
                "po_count",
                "item_count",
                "examples",
            ]
        )

    def make_examples(
        values: pd.Series,
    ) -> str:
        unique = list(
            dict.fromkeys(
                str(value).strip()
                for value in values
                if str(value).strip()
            )
        )

        return " | ".join(
            unique[:5]
        )

    result = (
        evidence
        .groupby(
            [
                "vendor_id",
                "classification",
            ],
            as_index=False,
        )
        .agg(
            po_count=(
                "po",
                "nunique",
            ),
            item_count=(
                "item_po",
                "size",
            ),
            examples=(
                "description",
                make_examples,
            ),
        )
    )

    return result.sort_values(
        [
            "vendor_id",
            "po_count",
            "item_count",
            "classification",
        ],
        ascending=[
            True,
            False,
            False,
            True,
        ],
    )


def build_vendor_rankings(
    aggregated: pd.DataFrame,
) -> dict[
    str,
    list[dict[str, object]],
]:
    result: dict[
        str,
        list[dict[str, object]],
    ] = defaultdict(
        list
    )

    for _, row in (
        aggregated.iterrows()
    ):
        result[
            row["vendor_id"]
        ].append(
            {
                "classification":
                    row[
                        "classification"
                    ],
                "po_count":
                    int(
                        row["po_count"]
                    ),
                "item_count":
                    int(
                        row["item_count"]
                    ),
                "examples":
                    row["examples"],
                "source":
                    "PO",
            }
        )

    return dict(
        result
    )


def build_unique_vendor_names(
    po: pd.DataFrame,
) -> dict[str, str]:
    mapping: dict[
        str,
        set[str],
    ] = defaultdict(
        set
    )

    for _, row in (
        po.iterrows()
    ):
        name = normalize_text(
            row["Nama Vendor"]
        )

        vendor_id = normalize_id(
            row["Vendor"]
        )

        if name and vendor_id:
            mapping[name].add(
                vendor_id
            )

    return {
        name: next(
            iter(
                vendor_ids
            )
        )
        for (
            name,
            vendor_ids,
        ) in mapping.items()
        if len(
            vendor_ids
        ) == 1
    }


def relation_status(
    classification: str,
    circle_support: set[str],
    taxonomy_support: set[str],
    circle_exists: bool,
    source: str,
) -> str:
    if source == "TAXONOMY_FALLBACK":
        return (
            "TAXONOMY_FALLBACK"
        )

    circle_match = (
        classification
        in circle_support
    )

    taxonomy_match = (
        classification
        in taxonomy_support
    )

    if (
        circle_match
        and taxonomy_match
    ):
        return (
            "CONFIRMED_CIRCLE_TAXONOMY"
        )

    if circle_match:
        return (
            "CONFIRMED_CIRCLE"
        )

    if taxonomy_match:
        return (
            "CONFIRMED_TAXONOMY"
        )

    if not circle_exists:
        return (
            "PO_ONLY_CIRCLE_MISSING"
        )

    return "PO_ONLY"


def main() -> None:
    V2_OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    # ======================================================
    # CONFIG / RULES
    # ======================================================

    vocabulary = load_vocabulary(
        VOCABULARY_V2_FILE
    )

    po_rules = load_po_rules(
        PO_RULES_V2_FILE,
        vocabulary,
    )

    context_rules = (
        load_context_rules(
            CONTEXT_RULES_V2_FILE,
            vocabulary,
        )
    )

    # ======================================================
    # LOAD AUDITED DATA CLEANSING
    # ======================================================

    cleansing = pd.read_excel(
        AUDIT_REVIEW_FILE,
        sheet_name=CLEANSING_SHEET,
        dtype=str,
        keep_default_na=False,
    )

    if (
        len(cleansing)
        != EXPECTED_VENDOR_ROWS
    ):
        raise RuntimeError(
            "Row guard failed: "
            f"expected "
            f"{EXPECTED_VENDOR_ROWS:,}, "
            f"got {len(cleansing):,}"
        )

    # ======================================================
    # LOAD DATA PO
    # ======================================================

    po = pd.read_excel(
        PO_FILE,
        sheet_name=PO_SHEET,
        header=PO_HEADER,
        dtype=str,
        keep_default_na=False,
    )

    require_columns(
        po,
        REQUIRED_PO_COLUMNS,
        "Data PO",
    )

    po["Doc.Date"] = (
        pd.to_datetime(
            po["Doc.Date"],
            errors="coerce",
            dayfirst=True,
        )
    )

    start_date = pd.Timestamp(
        PO_START_DATE
    )

    end_date = pd.Timestamp(
        PO_END_DATE
    )

    po = po.loc[
        po[
            "Doc.Date"
        ].between(
            start_date,
            end_date,
            inclusive="both",
        )
    ].copy()

    # ======================================================
    # CLASSIFY ACTUAL PO
    # ======================================================

    po_evidence = build_po_evidence(
        po,
        po_rules,
    )

    aggregated = (
        aggregate_po_evidence(
            po_evidence
        )
    )

    vendor_rankings = (
        build_vendor_rankings(
            aggregated
        )
    )

    unique_name_mapping = (
        build_unique_vendor_names(
            po
        )
    )

    # ======================================================
    # BUILD V2 PROPOSALS
    # ======================================================

    proposal_rows: list[
        dict[str, object]
    ] = []

    evidence_rows: list[
        dict[str, object]
    ] = []

    unresolved_rows: list[
        dict[str, object]
    ] = []

    existing_count = 0
    target_count = 0

    for df_index, row in (
        cleansing.iterrows()
    ):
        excel_row = (
            int(
                df_index
            )
            + 2
        )

        existing_final = str(
            row[
                FINAL_COLUMN
            ]
        ).strip()

        if existing_final:
            existing_count += 1

            proposal_rows.append(
                {
                    "excel_row":
                        excel_row,
                    "no_sap":
                        row[
                            SAP_COLUMN
                        ],
                    "vendor":
                        row[
                            VENDOR_COLUMN
                        ],
                    "status":
                        "EXISTING_PRESERVED",
                    "proposed_final":
                        existing_final,
                    "new_vocabulary":
                        "",
                }
            )

            continue

        target_count += 1

        vendor_id = normalize_id(
            row[
                SAP_COLUMN
            ]
        )

        vendor_name = normalize_text(
            row[
                VENDOR_COLUMN
            ]
        )

        if not vendor_id:
            vendor_id = (
                unique_name_mapping
                .get(
                    vendor_name,
                    "",
                )
            )

        circle_text = str(
            row.get(
                CIRCLE_COLUMN,
                "",
            )
        )

        circle_exists = bool(
            circle_text.strip()
        )

        taxonomy_text = " | ".join(
            [
                str(
                    row.get(
                        LEVEL_2_COLUMN,
                        "",
                    )
                ),
                str(
                    row.get(
                        LEVEL_3_COLUMN,
                        "",
                    )
                ),
            ]
        )

        circle_matches = (
            classify_context(
                circle_text,
                "CIRCLE",
                context_rules,
            )
        )

        taxonomy_matches = (
            classify_context(
                taxonomy_text,
                "TAXONOMY",
                context_rules,
            )
        )

        circle_support = set(
            circle_matches
        )

        taxonomy_support = set(
            taxonomy_matches
        )

        ranking = [
            dict(item)
            for item in (
                vendor_rankings
                .get(
                    vendor_id,
                    [],
                )
            )
        ]

        existing_labels = {
            item[
                "classification"
            ]
            for item in ranking
        }

        # ==================================================
        # SELECTIVE TAXONOMY FALLBACK
        #
        # Hanya context yang secara eksplisit
        # fallback_allowed=True.
        #
        # Electrical taxonomy tidak boleh membuat
        # ELEKTRIKAL sendiri.
        # ==================================================

        for (
            classification,
            context_rule,
        ) in taxonomy_matches.items():

            if (
                classification
                in existing_labels
            ):
                continue

            if not (
                context_rule
                .fallback_allowed
            ):
                continue

            ranking.append(
                {
                    "classification":
                        classification,
                    "po_count": 0,
                    "item_count": 0,
                    "examples":
                        (
                            "Taxonomy fallback: "
                            f"{context_rule.pattern_text}"
                        ),
                    "source":
                        "TAXONOMY_FALLBACK",
                }
            )

            existing_labels.add(
                classification
            )

        # ==================================================
        # SORT:
        # 1. PO count
        # 2. Item count
        # 3. fallback selalu di belakang
        # ==================================================

        ranking.sort(
            key=lambda item: (
                -int(
                    item[
                        "po_count"
                    ]
                ),
                -int(
                    item[
                        "item_count"
                    ]
                ),
                (
                    1
                    if item[
                        "source"
                    ]
                    == "TAXONOMY_FALLBACK"
                    else 0
                ),
                item[
                    "classification"
                ],
            )
        )

        labels = [
            item[
                "classification"
            ]
            for item in ranking
        ]

        proposed_final = (
            ", ".join(
                labels
            )
        )

        new_labels = [
            label
            for label in labels
            if (
                vocabulary.get(
                    label
                )
                == "NEW"
            )
        ]

        status = (
            "PROPOSED_V2"
            if proposed_final
            else "UNRESOLVED"
        )

        proposal_rows.append(
            {
                "excel_row":
                    excel_row,
                "no_sap":
                    row[
                        SAP_COLUMN
                    ],
                "vendor":
                    row[
                        VENDOR_COLUMN
                    ],
                "status":
                    status,
                "proposed_final":
                    proposed_final,
                "new_vocabulary":
                    ", ".join(
                        new_labels
                    ),
            }
        )

        if not proposed_final:
            unresolved_rows.append(
                {
                    "excel_row":
                        excel_row,
                    "no_sap":
                        row[
                            SAP_COLUMN
                        ],
                    "vendor":
                        row[
                            VENDOR_COLUMN
                        ],
                    "circle":
                        circle_text,
                    "items":
                        row.get(
                            PO_ITEM_COLUMN,
                            "",
                        ),
                    "level_2":
                        row.get(
                            LEVEL_2_COLUMN,
                            "",
                        ),
                    "level_3":
                        row.get(
                            LEVEL_3_COLUMN,
                            "",
                        ),
                }
            )

        for rank_number, item in enumerate(
            ranking,
            start=1,
        ):
            classification = (
                item[
                    "classification"
                ]
            )

            relation = relation_status(
                classification=(
                    classification
                ),
                circle_support=(
                    circle_support
                ),
                taxonomy_support=(
                    taxonomy_support
                ),
                circle_exists=(
                    circle_exists
                ),
                source=(
                    item[
                        "source"
                    ]
                ),
            )

            evidence_rows.append(
                {
                    "excel_row":
                        excel_row,
                    "no_sap":
                        row[
                            SAP_COLUMN
                        ],
                    "vendor":
                        row[
                            VENDOR_COLUMN
                        ],
                    "rank":
                        rank_number,
                    "classification":
                        classification,
                    "vocabulary_status":
                        vocabulary.get(
                            classification,
                            "",
                        ),
                    "relation":
                        relation,
                    "circle_support":
                        (
                            classification
                            in circle_support
                        ),
                    "taxonomy_support":
                        (
                            classification
                            in taxonomy_support
                        ),
                    "distinct_po_count":
                        item[
                            "po_count"
                        ],
                    "item_count":
                        item[
                            "item_count"
                        ],
                    "source":
                        item[
                            "source"
                        ],
                    "examples":
                        item[
                            "examples"
                        ],
                }
            )

    proposal_df = pd.DataFrame(
        proposal_rows
    )

    evidence_df = pd.DataFrame(
        evidence_rows
    )

    unresolved_df = pd.DataFrame(
        unresolved_rows
    )

    # ======================================================
    # NEW VOCABULARY USAGE
    # ======================================================

    if evidence_df.empty:
        new_usage = pd.DataFrame(
            columns=[
                "classification",
                "vendor_count",
                "evidence_rows",
            ]
        )
    else:
        new_usage = (
            evidence_df.loc[
                evidence_df[
                    "vocabulary_status"
                ]
                .eq(
                    "NEW"
                )
            ]
            .groupby(
                "classification",
                as_index=False,
            )
            .agg(
                vendor_count=(
                    "excel_row",
                    "nunique",
                ),
                evidence_rows=(
                    "excel_row",
                    "size",
                ),
            )
            .sort_values(
                "vendor_count",
                ascending=False,
            )
        )

    # ======================================================
    # EXPORT CSV
    # ======================================================

    proposal_df.to_csv(
        V2_OUTPUT_DIR
        / "classification_proposals_v2.csv",
        index=False,
        encoding="utf-8-sig",
    )

    evidence_df.to_csv(
        V2_OUTPUT_DIR
        / "classification_evidence_v2.csv",
        index=False,
        encoding="utf-8-sig",
    )

    unresolved_df.to_csv(
        V2_OUTPUT_DIR
        / "classification_unresolved_v2.csv",
        index=False,
        encoding="utf-8-sig",
    )

    new_usage.to_csv(
        V2_OUTPUT_DIR
        / "new_vocabulary_usage.csv",
        index=False,
        encoding="utf-8-sig",
    )

    # ======================================================
    # BUILD WORKBOOK
    # ======================================================

    output_file = (
        V2_PREVIEW_OUTPUT_FILE
        if WRITE_V2_PREVIEW_ONLY
        else V2_FINAL_OUTPUT_FILE
    )

    shutil.copy2(
        AUDIT_REVIEW_FILE,
        output_file,
    )

    workbook = load_workbook(
        output_file
    )

    worksheet = workbook[
        CLEANSING_SHEET
    ]

    final_column_index = (
        find_header_column(
            worksheet,
            FINAL_COLUMN,
        )
    )

    proposal_lookup = (
        proposal_df
        .set_index(
            "excel_row"
        )
    )

    new_fill = PatternFill(
        fill_type="solid",
        fgColor=(
            NEW_VOCABULARY_FILL
        ),
    )

    unresolved_fill = (
        PatternFill(
            fill_type="solid",
            fgColor=(
                UNRESOLVED_CLASSIFICATION_FILL
            ),
        )
    )

    # ======================================================
    # WRITE ONLY ORIGINAL BLANK T
    # ======================================================

    for (
        excel_row,
        proposal,
    ) in proposal_lookup.iterrows():

        cell = worksheet.cell(
            row=int(
                excel_row
            ),
            column=(
                final_column_index
            ),
        )

        if str(
            cell.value or ""
        ).strip():
            continue

        proposed = str(
            proposal[
                "proposed_final"
            ]
        ).strip()

        if proposed:
            cell.value = proposed

            new_vocab = str(
                proposal[
                    "new_vocabulary"
                ]
            ).strip()

            if new_vocab:
                cell.fill = (
                    new_fill
                )
                cell.font = Font(
                    bold=True
                )

        else:
            cell.fill = (
                unresolved_fill
            )
            cell.font = Font(
                bold=True
            )

    # ======================================================
    # EVIDENCE SHEET
    # ======================================================

    sheet_name = (
        "Classification Evidence V2"
    )

    if (
        sheet_name
        in workbook.sheetnames
    ):
        del workbook[
            sheet_name
        ]

    evidence_ws = (
        workbook.create_sheet(
            sheet_name
        )
    )

    headers = [
        "Excel Row",
        "NO SAP",
        "Vendor",
        "Rank",
        "Classification",
        "Vocabulary Status",
        "Relation",
        "Circle Support",
        "Taxonomy Support",
        "Distinct PO Count",
        "Item Count",
        "Source",
        "Examples",
    ]

    evidence_ws.append(
        headers
    )

    for cell in evidence_ws[1]:
        cell.font = Font(
            bold=True
        )

    if not evidence_df.empty:
        for _, row in (
            evidence_df.iterrows()
        ):
            evidence_ws.append(
                [
                    row[
                        "excel_row"
                    ],
                    row[
                        "no_sap"
                    ],
                    row[
                        "vendor"
                    ],
                    row[
                        "rank"
                    ],
                    row[
                        "classification"
                    ],
                    row[
                        "vocabulary_status"
                    ],
                    row[
                        "relation"
                    ],
                    row[
                        "circle_support"
                    ],
                    row[
                        "taxonomy_support"
                    ],
                    row[
                        "distinct_po_count"
                    ],
                    row[
                        "item_count"
                    ],
                    row[
                        "source"
                    ],
                    row[
                        "examples"
                    ],
                ]
            )

    evidence_ws.freeze_panes = (
        "A2"
    )

    evidence_ws.auto_filter.ref = (
        evidence_ws.dimensions
    )

    # ======================================================
    # NEW VOCABULARY SHEET
    # ======================================================

    sheet_name = (
        "New Vocabulary Usage"
    )

    if (
        sheet_name
        in workbook.sheetnames
    ):
        del workbook[
            sheet_name
        ]

    new_ws = (
        workbook.create_sheet(
            sheet_name
        )
    )

    new_ws.append(
        [
            "Classification",
            "Vendor Count",
            "Evidence Rows",
        ]
    )

    for cell in new_ws[1]:
        cell.font = Font(
            bold=True
        )

    if not new_usage.empty:
        for _, row in (
            new_usage.iterrows()
        ):
            new_ws.append(
                [
                    row[
                        "classification"
                    ],
                    row[
                        "vendor_count"
                    ],
                    row[
                        "evidence_rows"
                    ],
                ]
            )

    # ======================================================
    # UNRESOLVED SHEET
    # ======================================================

    sheet_name = (
        "Classification Review V2"
    )

    if (
        sheet_name
        in workbook.sheetnames
    ):
        del workbook[
            sheet_name
        ]

    review_ws = (
        workbook.create_sheet(
            sheet_name
        )
    )

    review_ws.append(
        [
            "Excel Row",
            "NO SAP",
            "Vendor",
            "Circle",
            "Item Pekerjaan",
            "Level 2",
            "Level 3",
        ]
    )

    for cell in review_ws[1]:
        cell.font = Font(
            bold=True
        )

    if not unresolved_df.empty:
        for _, row in (
            unresolved_df.iterrows()
        ):
            review_ws.append(
                [
                    row[
                        "excel_row"
                    ],
                    row[
                        "no_sap"
                    ],
                    row[
                        "vendor"
                    ],
                    row[
                        "circle"
                    ],
                    row[
                        "items"
                    ],
                    row[
                        "level_2"
                    ],
                    row[
                        "level_3"
                    ],
                ]
            )

    workbook.save(
        output_file
    )

    # ======================================================
    # VERIFY
    # ======================================================

    verify = pd.read_excel(
        output_file,
        sheet_name=(
            CLEANSING_SHEET
        ),
        dtype=str,
        keep_default_na=False,
    )

    if (
        len(verify)
        != EXPECTED_VENDOR_ROWS
    ):
        raise RuntimeError(
            "OUTPUT ROW GUARD FAILED"
        )

    remaining_blank = int(
        verify[
            FINAL_COLUMN
        ]
        .astype(str)
        .str.strip()
        .eq("")
        .sum()
    )

    print()
    print("=" * 72)
    print(
        "CLASSIFICATION V2"
    )
    print("=" * 72)

    print(
        f"Vendor rows             : "
        f"{len(cleansing):,}"
    )

    print(
        f"Existing preserved      : "
        f"{existing_count:,}"
    )

    print(
        f"Blank targets           : "
        f"{target_count:,}"
    )

    print(
        f"V2 classified           : "
        f"{target_count - len(unresolved_df):,}"
    )

    print(
        f"Unresolved              : "
        f"{len(unresolved_df):,}"
    )

    print(
        f"Remaining blank T       : "
        f"{remaining_blank:,}"
    )

    print(
        f"New vocabulary used     : "
        f"{len(new_usage):,}"
    )

    print(
        f"Output                  : "
        f"{output_file}"
    )

    print("=" * 72)


if __name__ == "__main__":
    main()