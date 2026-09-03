from __future__ import annotations

import shutil
import sys
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = (
    Path(__file__)
    .resolve()
    .parents[1]
)

if str(ROOT) not in sys.path:
    sys.path.insert(
        0,
        str(ROOT),
    )


from src.classifier_v2_1 import (
    classify_context,
    classify_primary_text,
    load_context_rules,
    load_exclusion_rules,
    load_primary_rules,
    load_vocabulary,
    normalize_id,
    normalize_text,
    split_work_items,
)

from src.config import (
    AUDIT_REVIEW_FILE,
    CLASSIFICATION_EXCLUSIONS_V2_FILE,
    CLEANSING_SHEET,
    CONTEXT_RULES_V2_FILE,
    EXPECTED_VENDOR_ROWS,
    NEW_VOCABULARY_FILL,
    PO_END_DATE,
    PO_FILE,
    PO_HEADER,
    PO_RULES_V2_FILE,
    PO_SHEET,
    PO_START_DATE,
    REQUIRED_PO_COLUMNS,
    UNRESOLVED_CLASSIFICATION_FILL,
    V2_1_FINAL_OUTPUT_FILE,
    V2_1_OUTPUT_DIR,
    V2_1_PREVIEW_OUTPUT_FILE,
    V2_BASELINE_PREVIEW_FILE,
    VOCABULARY_V2_FILE,
    WRITE_V2_1_PREVIEW_ONLY,
)


FINAL_COLUMN = "Klasifikasi Final"
SAP_COLUMN = "NO SAP"
VENDOR_COLUMN = "Nama Rekanan"
CIRCLE_COLUMN = "Klasifikasi Circle"

PO_ITEM_COLUMN = (
    "Item Pekerjaan Berdasarkan PO"
)

LEVEL_1_COLUMN = (
    "Kelompok Klasifikasi\n"
    "Level 1 (Klasifikasi PO)"
)

LEVEL_2_COLUMN = (
    "Kelompok Klasifikasi\n"
    "Level 2 (Klasifikasi PO)"
)

LEVEL_3_COLUMN = (
    "Kelompok Klasifikasi\n"
    "Level 3 (Klasifikasi PO)"
)


def require_columns(
    df: pd.DataFrame,
    required: list[str],
    source_name: str,
) -> None:
    missing = [
        column
        for column in required
        if column not in df.columns
    ]

    if missing:
        raise ValueError(
            f"{source_name}: "
            f"missing columns: {missing}"
        )


def find_header_column(
    worksheet,
    header_name: str,
) -> int:
    for cell in worksheet[1]:
        if (
            str(
                cell.value
            ).strip()
            == header_name
        ):
            return int(
                cell.column
            )

    raise ValueError(
        f"Header not found: "
        f"{header_name}"
    )


def build_unique_vendor_names(
    po: pd.DataFrame,
) -> dict[str, str]:
    mapping: dict[
        str,
        set[str],
    ] = defaultdict(
        set
    )

    for _, row in po.iterrows():
        name = normalize_text(
            row["Nama Vendor"]
        )

        vendor_id = normalize_id(
            row["Vendor"]
        )

        if (
            name
            and vendor_id
        ):
            mapping[
                name
            ].add(
                vendor_id
            )

    return {
        name: next(
            iter(vendor_ids)
        )
        for (
            name,
            vendor_ids,
        ) in mapping.items()
        if len(
            vendor_ids
        ) == 1
    }


def build_raw_po_evidence(
    po: pd.DataFrame,
    primary_rules,
    exclusion_rules,
) -> pd.DataFrame:
    rows: list[
        dict[str, object]
    ] = []

    for _, row in po.iterrows():
        vendor_id = normalize_id(
            row["Vendor"]
        )

        po_number = normalize_id(
            row["PO"]
        )

        item_po = normalize_id(
            row["Item.PO"]
        )

        description = str(
            row["Deskripsi"]
        ).strip()

        matches = (
            classify_primary_text(
                description,
                primary_rules,
                exclusion_rules,
            )
        )

        for (
            classification,
            rule,
        ) in matches.items():
            rows.append(
                {
                    "vendor_id":
                        vendor_id,
                    "po":
                        po_number,
                    "item_po":
                        item_po,
                    "description":
                        description,
                    "classification":
                        classification,
                    "priority":
                        rule.priority,
                    "rule_pattern":
                        rule.pattern_text,
                }
            )

    if not rows:
        return pd.DataFrame(
            columns=[
                "vendor_id",
                "po",
                "item_po",
                "description",
                "classification",
                "priority",
                "rule_pattern",
            ]
        )

    return pd.DataFrame(
        rows
    )


def aggregate_raw_po(
    evidence: pd.DataFrame,
) -> pd.DataFrame:
    if evidence.empty:
        return pd.DataFrame(
            columns=[
                "vendor_id",
                "classification",
                "po_count",
                "po_item_count",
                "po_examples",
            ]
        )

    def make_examples(
        values: pd.Series,
    ) -> str:
        unique = list(
            dict.fromkeys(
                str(value).strip()
                for value in values
                if str(value).strip()
            )
        )

        return " | ".join(
            unique[:5]
        )

    result = (
        evidence
        .groupby(
            [
                "vendor_id",
                "classification",
            ],
            as_index=False,
        )
        .agg(
            po_count=(
                "po",
                "nunique",
            ),
            po_item_count=(
                "description",
                "size",
            ),
            po_examples=(
                "description",
                make_examples,
            ),
        )
    )

    return result


def build_raw_po_lookup(
    aggregated: pd.DataFrame,
) -> dict[
    str,
    dict[
        str,
        dict[str, object],
    ],
]:
    result: dict[
        str,
        dict[
            str,
            dict[str, object],
        ],
    ] = defaultdict(
        dict
    )

    for _, row in (
        aggregated.iterrows()
    ):
        result[
            row["vendor_id"]
        ][
            row["classification"]
        ] = {
            "classification":
                row["classification"],
            "po_count":
                int(
                    row["po_count"]
                ),
            "po_item_count":
                int(
                    row["po_item_count"]
                ),
            "po_examples":
                row["po_examples"],
        }

    return dict(
        result
    )


def classify_u_items(
    value: object,
    primary_rules,
    exclusion_rules,
) -> dict[
    str,
    dict[str, object],
]:
    result: dict[
        str,
        dict[str, object],
    ] = {}

    items = split_work_items(
        value
    )

    for item in items:
        matches = (
            classify_primary_text(
                item,
                primary_rules,
                exclusion_rules,
            )
        )

        for classification in (
            matches.keys()
        ):
            entry = result.setdefault(
                classification,
                {
                    "u_item_count": 0,
                    "u_examples": [],
                },
            )

            entry[
                "u_item_count"
            ] += 1

            if (
                item
                not in entry[
                    "u_examples"
                ]
            ):
                entry[
                    "u_examples"
                ].append(
                    item
                )

    for value_dict in (
        result.values()
    ):
        value_dict[
            "u_examples"
        ] = " | ".join(
            value_dict[
                "u_examples"
            ][:5]
        )

    return result


def combine_primary_evidence(
    vendor_id: str,
    u_evidence: dict[
        str,
        dict[str, object],
    ],
    raw_lookup: dict[
        str,
        dict[
            str,
            dict[str, object],
        ],
    ],
) -> dict[
    str,
    dict[str, object],
]:
    combined: dict[
        str,
        dict[str, object],
    ] = {}

    raw_vendor = (
        raw_lookup.get(
            vendor_id,
            {},
        )
    )

    for (
        classification,
        raw_item,
    ) in raw_vendor.items():
        combined[
            classification
        ] = {
            "classification":
                classification,
            "po_count":
                int(
                    raw_item[
                        "po_count"
                    ]
                ),
            "po_item_count":
                int(
                    raw_item[
                        "po_item_count"
                    ]
                ),
            "u_item_count":
                0,
            "po_examples":
                raw_item[
                    "po_examples"
                ],
            "u_examples":
                "",
            "source":
                "PO",
        }

    for (
        classification,
        u_item,
    ) in u_evidence.items():
        if (
            classification
            in combined
        ):
            combined[
                classification
            ][
                "u_item_count"
            ] = int(
                u_item[
                    "u_item_count"
                ]
            )

            combined[
                classification
            ][
                "u_examples"
            ] = u_item[
                "u_examples"
            ]

            combined[
                classification
            ][
                "source"
            ] = "PO+ITEM_U"

        else:
            combined[
                classification
            ] = {
                "classification":
                    classification,
                "po_count": 0,
                "po_item_count": 0,
                "u_item_count":
                    int(
                        u_item[
                            "u_item_count"
                        ]
                    ),
                "po_examples": "",
                "u_examples":
                    u_item[
                        "u_examples"
                    ],
                "source":
                    "ITEM_U",
            }

    return combined


def add_taxonomy_fallback(
    combined: dict[
        str,
        dict[str, object],
    ],
    taxonomy_matches,
    has_primary_text: bool,
) -> None:
    # Taxonomy hanya boleh menghasilkan final
    # ketika primary evidence belum menghasilkan
    # classification sama sekali.
    if combined:
        return

    if not has_primary_text:
        return

    for (
        classification,
        context_rule,
    ) in taxonomy_matches.items():
        if not (
            context_rule
            .fallback_allowed
        ):
            continue

        combined[
            classification
        ] = {
            "classification":
                classification,
            "po_count": 0,
            "po_item_count": 0,
            "u_item_count": 0,
            "po_examples": "",
            "u_examples": "",
            "source":
                "TAXONOMY_FALLBACK",
        }


def sort_classifications(
    combined: dict[
        str,
        dict[str, object],
    ],
) -> list[
    dict[str, object]
]:
    values = list(
        combined.values()
    )

    source_order = {
        "PO+ITEM_U": 0,
        "PO": 1,
        "ITEM_U": 2,
        "TAXONOMY_FALLBACK": 3,
    }

    values.sort(
        key=lambda item: (
            -int(
                item[
                    "po_count"
                ]
            ),
            -(
                int(
                    item[
                        "po_item_count"
                    ]
                )
                + int(
                    item[
                        "u_item_count"
                    ]
                )
            ),
            source_order.get(
                item[
                    "source"
                ],
                99,
            ),
            item[
                "classification"
            ],
        )
    )

    return values


def relation_status(
    classification: str,
    source: str,
    circle_support: set[str],
    taxonomy_support: set[str],
    circle_exists: bool,
) -> str:
    circle_match = (
        classification
        in circle_support
    )

    taxonomy_match = (
        classification
        in taxonomy_support
    )

    if (
        circle_match
        and taxonomy_match
    ):
        return (
            "CONFIRMED_CIRCLE_TAXONOMY"
        )

    if circle_match:
        return (
            "CONFIRMED_CIRCLE"
        )

    if taxonomy_match:
        return (
            "CONFIRMED_TAXONOMY"
        )

    if (
        source
        == "TAXONOMY_FALLBACK"
    ):
        return (
            "TAXONOMY_FALLBACK"
        )

    if not circle_exists:
        return (
            "ACTUAL_WORK_CIRCLE_MISSING"
        )

    return (
        "ACTUAL_WORK_NOT_DECLARED"
    )


def append_dataframe_sheet(
    workbook,
    sheet_name: str,
    dataframe: pd.DataFrame,
) -> None:
    if (
        sheet_name
        in workbook.sheetnames
    ):
        del workbook[
            sheet_name
        ]

    ws = workbook.create_sheet(
        sheet_name
    )

    if dataframe.empty:
        ws.append(
            ["No data"]
        )
        return

    ws.append(
        list(
            dataframe.columns
        )
    )

    for cell in ws[1]:
        cell.font = Font(
            bold=True
        )

    for row in (
        dataframe.itertuples(
            index=False,
            name=None,
        )
    ):
        ws.append(
            list(row)
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        ws.dimensions
    )


def build_v2_comparison(
    cleansing: pd.DataFrame,
    proposal_df: pd.DataFrame,
) -> pd.DataFrame:
    if not (
        V2_BASELINE_PREVIEW_FILE
        .exists()
    ):
        return pd.DataFrame(
            columns=[
                "excel_row",
                "no_sap",
                "vendor",
                "v2",
                "v2_1",
                "changed",
            ]
        )

    baseline = pd.read_excel(
        V2_BASELINE_PREVIEW_FILE,
        sheet_name=CLEANSING_SHEET,
        dtype=str,
        keep_default_na=False,
    )

    if (
        len(baseline)
        != len(cleansing)
    ):
        return pd.DataFrame(
            columns=[
                "excel_row",
                "no_sap",
                "vendor",
                "v2",
                "v2_1",
                "changed",
            ]
        )

    proposal_lookup = (
        proposal_df.set_index(
            "excel_row"
        )
    )

    rows = []

    for df_index, original_row in (
        cleansing.iterrows()
    ):
        original_final = str(
            original_row[
                FINAL_COLUMN
            ]
        ).strip()

        if original_final:
            continue

        excel_row = (
            int(df_index)
            + 2
        )

        if (
            excel_row
            not in proposal_lookup.index
        ):
            continue

        v2_value = str(
            baseline.iloc[
                int(df_index)
            ][
                FINAL_COLUMN
            ]
        ).strip()

        v2_1_value = str(
            proposal_lookup.loc[
                excel_row,
                "proposed_final",
            ]
        ).strip()

        if (
            v2_value
            == v2_1_value
        ):
            continue

        rows.append(
            {
                "excel_row":
                    excel_row,
                "no_sap":
                    original_row[
                        SAP_COLUMN
                    ],
                "vendor":
                    original_row[
                        VENDOR_COLUMN
                    ],
                "v2":
                    v2_value,
                "v2_1":
                    v2_1_value,
                "changed":
                    True,
            }
        )

    return pd.DataFrame(
        rows
    )


def main() -> None:
    V2_1_OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    # ======================================================
    # RULES
    # ======================================================

    vocabulary = load_vocabulary(
        VOCABULARY_V2_FILE
    )

    primary_rules = (
        load_primary_rules(
            PO_RULES_V2_FILE,
            vocabulary,
        )
    )

    context_rules = (
        load_context_rules(
            CONTEXT_RULES_V2_FILE,
            vocabulary,
        )
    )

    exclusion_rules = (
        load_exclusion_rules(
            CLASSIFICATION_EXCLUSIONS_V2_FILE,
            vocabulary,
        )
    )

    # ======================================================
    # DATA CLEANSING
    # ======================================================

    cleansing = pd.read_excel(
        AUDIT_REVIEW_FILE,
        sheet_name=CLEANSING_SHEET,
        dtype=str,
        keep_default_na=False,
    )

    if (
        len(cleansing)
        != EXPECTED_VENDOR_ROWS
    ):
        raise RuntimeError(
            "INPUT ROW GUARD FAILED: "
            f"expected "
            f"{EXPECTED_VENDOR_ROWS:,}, "
            f"got "
            f"{len(cleansing):,}"
        )

    require_columns(
        cleansing,
        [
            SAP_COLUMN,
            VENDOR_COLUMN,
            CIRCLE_COLUMN,
            FINAL_COLUMN,
            PO_ITEM_COLUMN,
            LEVEL_1_COLUMN,
            LEVEL_2_COLUMN,
            LEVEL_3_COLUMN,
        ],
        "Data Cleansing",
    )

    # ======================================================
    # RAW PO
    # ======================================================

    po = pd.read_excel(
        PO_FILE,
        sheet_name=PO_SHEET,
        header=PO_HEADER,
        dtype=str,
        keep_default_na=False,
    )

    require_columns(
        po,
        REQUIRED_PO_COLUMNS,
        "Data PO",
    )

    po["Doc.Date"] = (
        pd.to_datetime(
            po["Doc.Date"],
            errors="coerce",
            dayfirst=True,
        )
    )

    start_date = pd.Timestamp(
        PO_START_DATE
    )

    end_date = pd.Timestamp(
        PO_END_DATE
    )

    po = po.loc[
        po[
            "Doc.Date"
        ].between(
            start_date,
            end_date,
            inclusive="both",
        )
    ].copy()

    raw_evidence = (
        build_raw_po_evidence(
            po,
            primary_rules,
            exclusion_rules,
        )
    )

    raw_aggregated = (
        aggregate_raw_po(
            raw_evidence
        )
    )

    raw_lookup = (
        build_raw_po_lookup(
            raw_aggregated
        )
    )

    unique_vendor_names = (
        build_unique_vendor_names(
            po
        )
    )

    # ======================================================
    # PROPOSALS
    # ======================================================

    proposal_rows = []
    evidence_rows = []
    unresolved_rows = []

    existing_count = 0
    target_count = 0

    source_counter = (
        defaultdict(int)
    )

    for (
        df_index,
        row,
    ) in cleansing.iterrows():

        excel_row = (
            int(df_index)
            + 2
        )

        existing_final = str(
            row[
                FINAL_COLUMN
            ]
        ).strip()

        if existing_final:
            existing_count += 1

            proposal_rows.append(
                {
                    "excel_row":
                        excel_row,
                    "no_sap":
                        row[
                            SAP_COLUMN
                        ],
                    "vendor":
                        row[
                            VENDOR_COLUMN
                        ],
                    "status":
                        "EXISTING_PRESERVED",
                    "proposed_final":
                        existing_final,
                    "new_vocabulary":
                        "",
                    "vendor_match_method":
                        "NOT_REQUIRED",
                }
            )

            continue

        target_count += 1

        vendor_id = normalize_id(
            row[
                SAP_COLUMN
            ]
        )

        vendor_match_method = (
            "NO_SAP"
            if vendor_id
            else ""
        )

        if not vendor_id:
            vendor_name = (
                normalize_text(
                    row[
                        VENDOR_COLUMN
                    ]
                )
            )

            vendor_id = (
                unique_vendor_names.get(
                    vendor_name,
                    "",
                )
            )

            if vendor_id:
                vendor_match_method = (
                    "UNIQUE_VENDOR_NAME"
                )
            else:
                vendor_match_method = (
                    "NO_VENDOR_MATCH"
                )

        u_text = str(
            row[
                PO_ITEM_COLUMN
            ]
        )

        u_evidence = (
            classify_u_items(
                u_text,
                primary_rules,
                exclusion_rules,
            )
        )

        combined = (
            combine_primary_evidence(
                vendor_id,
                u_evidence,
                raw_lookup,
            )
        )

        circle_text = str(
            row[
                CIRCLE_COLUMN
            ]
        )

        circle_exists = bool(
            circle_text.strip()
        )

        taxonomy_text = " | ".join(
            [
                str(
                    row[
                        LEVEL_1_COLUMN
                    ]
                ),
                str(
                    row[
                        LEVEL_2_COLUMN
                    ]
                ),
                str(
                    row[
                        LEVEL_3_COLUMN
                    ]
                ),
            ]
        )

        circle_matches = (
            classify_context(
                circle_text,
                "CIRCLE",
                context_rules,
            )
        )

        taxonomy_matches = (
            classify_context(
                taxonomy_text,
                "TAXONOMY",
                context_rules,
            )
        )

        has_primary_text = bool(
            u_text.strip()
        ) or bool(
            vendor_id
            and vendor_id
            in raw_lookup
        )

        add_taxonomy_fallback(
            combined,
            taxonomy_matches,
            has_primary_text,
        )

        ranking = (
            sort_classifications(
                combined
            )
        )

        labels = [
            item[
                "classification"
            ]
            for item in ranking
        ]

        proposed_final = (
            ", ".join(
                labels
            )
        )

        new_labels = [
            classification
            for classification in labels
            if (
                vocabulary.get(
                    classification
                )
                == "NEW"
            )
        ]

        status = (
            "PROPOSED_V2_1"
            if proposed_final
            else "UNRESOLVED"
        )

        proposal_rows.append(
            {
                "excel_row":
                    excel_row,
                "no_sap":
                    row[
                        SAP_COLUMN
                    ],
                "vendor":
                    row[
                        VENDOR_COLUMN
                    ],
                "status":
                    status,
                "proposed_final":
                    proposed_final,
                "new_vocabulary":
                    ", ".join(
                        new_labels
                    ),
                "vendor_match_method":
                    vendor_match_method,
            }
        )

        if not proposed_final:
            unresolved_rows.append(
                {
                    "excel_row":
                        excel_row,
                    "no_sap":
                        row[
                            SAP_COLUMN
                        ],
                    "vendor":
                        row[
                            VENDOR_COLUMN
                        ],
                    "vendor_match_method":
                        vendor_match_method,
                    "circle":
                        circle_text,
                    "item_pekerjaan":
                        u_text,
                    "level_1":
                        row[
                            LEVEL_1_COLUMN
                        ],
                    "level_2":
                        row[
                            LEVEL_2_COLUMN
                        ],
                    "level_3":
                        row[
                            LEVEL_3_COLUMN
                        ],
                    "reason":
                        (
                            "Tidak ada primary rule "
                            "atau taxonomy fallback "
                            "yang cukup kuat"
                        ),
                }
            )

        circle_support = set(
            circle_matches.keys()
        )

        taxonomy_support = set(
            taxonomy_matches.keys()
        )

        for (
            rank_number,
            item,
        ) in enumerate(
            ranking,
            start=1,
        ):
            classification = (
                item[
                    "classification"
                ]
            )

            source = (
                item[
                    "source"
                ]
            )

            source_counter[
                source
            ] += 1

            relation = (
                relation_status(
                    classification=(
                        classification
                    ),
                    source=source,
                    circle_support=(
                        circle_support
                    ),
                    taxonomy_support=(
                        taxonomy_support
                    ),
                    circle_exists=(
                        circle_exists
                    ),
                )
            )

            evidence_rows.append(
                {
                    "excel_row":
                        excel_row,
                    "no_sap":
                        row[
                            SAP_COLUMN
                        ],
                    "vendor":
                        row[
                            VENDOR_COLUMN
                        ],
                    "rank":
                        rank_number,
                    "classification":
                        classification,
                    "vocabulary_status":
                        vocabulary.get(
                            classification,
                            "",
                        ),
                    "source":
                        source,
                    "relation":
                        relation,
                    "circle_support":
                        classification
                        in circle_support,
                    "taxonomy_support":
                        classification
                        in taxonomy_support,
                    "distinct_po_count":
                        item[
                            "po_count"
                        ],
                    "raw_po_item_count":
                        item[
                            "po_item_count"
                        ],
                    "column_u_item_count":
                        item[
                            "u_item_count"
                        ],
                    "raw_po_examples":
                        item[
                            "po_examples"
                        ],
                    "column_u_examples":
                        item[
                            "u_examples"
                        ],
                    "circle":
                        circle_text,
                    "taxonomy":
                        taxonomy_text,
                }
            )

    proposal_df = (
        pd.DataFrame(
            proposal_rows
        )
    )

    evidence_df = (
        pd.DataFrame(
            evidence_rows
        )
    )

    unresolved_df = (
        pd.DataFrame(
            unresolved_rows
        )
    )

    # ======================================================
    # NEW VOCABULARY USAGE
    # ======================================================

    if evidence_df.empty:
        new_usage_df = (
            pd.DataFrame(
                columns=[
                    "classification",
                    "vendor_count",
                    "evidence_rows",
                    "distinct_po_total",
                    "column_u_item_total",
                ]
            )
        )
    else:
        new_evidence = (
            evidence_df.loc[
                evidence_df[
                    "vocabulary_status"
                ]
                .eq(
                    "NEW"
                )
            ]
            .copy()
        )

        if new_evidence.empty:
            new_usage_df = (
                pd.DataFrame(
                    columns=[
                        "classification",
                        "vendor_count",
                        "evidence_rows",
                        "distinct_po_total",
                        "column_u_item_total",
                    ]
                )
            )
        else:
            new_usage_df = (
                new_evidence
                .groupby(
                    "classification",
                    as_index=False,
                )
                .agg(
                    vendor_count=(
                        "excel_row",
                        "nunique",
                    ),
                    evidence_rows=(
                        "excel_row",
                        "size",
                    ),
                    distinct_po_total=(
                        "distinct_po_count",
                        "sum",
                    ),
                    column_u_item_total=(
                        "column_u_item_count",
                        "sum",
                    ),
                )
                .sort_values(
                    [
                        "vendor_count",
                        "classification",
                    ],
                    ascending=[
                        False,
                        True,
                    ],
                )
            )

    # ======================================================
    # V2 VS V2.1
    # ======================================================

    comparison_df = (
        build_v2_comparison(
            cleansing,
            proposal_df,
        )
    )

    # ======================================================
    # CSV OUTPUT
    # ======================================================

    proposal_df.to_csv(
        V2_1_OUTPUT_DIR
        / "classification_proposals_v2_1.csv",
        index=False,
        encoding="utf-8-sig",
    )

    evidence_df.to_csv(
        V2_1_OUTPUT_DIR
        / "classification_evidence_v2_1.csv",
        index=False,
        encoding="utf-8-sig",
    )

    unresolved_df.to_csv(
        V2_1_OUTPUT_DIR
        / "classification_unresolved_v2_1.csv",
        index=False,
        encoding="utf-8-sig",
    )

    new_usage_df.to_csv(
        V2_1_OUTPUT_DIR
        / "new_vocabulary_usage_v2_1.csv",
        index=False,
        encoding="utf-8-sig",
    )

    comparison_df.to_csv(
        V2_1_OUTPUT_DIR
        / "comparison_v2_vs_v2_1.csv",
        index=False,
        encoding="utf-8-sig",
    )

    # ======================================================
    # WORKBOOK
    # ======================================================

    output_file = (
        V2_1_PREVIEW_OUTPUT_FILE
        if WRITE_V2_1_PREVIEW_ONLY
        else V2_1_FINAL_OUTPUT_FILE
    )

    shutil.copy2(
        AUDIT_REVIEW_FILE,
        output_file,
    )

    workbook = load_workbook(
        output_file
    )

    worksheet = workbook[
        CLEANSING_SHEET
    ]

    final_column_index = (
        find_header_column(
            worksheet,
            FINAL_COLUMN,
        )
    )

    proposal_lookup = (
        proposal_df.set_index(
            "excel_row"
        )
    )

    new_fill = PatternFill(
        fill_type="solid",
        fgColor=(
            NEW_VOCABULARY_FILL
        ),
    )

    unresolved_fill = (
        PatternFill(
            fill_type="solid",
            fgColor=(
                UNRESOLVED_CLASSIFICATION_FILL
            ),
        )
    )

    for (
        excel_row,
        proposal,
    ) in proposal_lookup.iterrows():

        cell = worksheet.cell(
            row=int(
                excel_row
            ),
            column=(
                final_column_index
            ),
        )

        # Existing Final T tidak boleh diubah.
        if str(
            cell.value or ""
        ).strip():
            continue

        proposed = str(
            proposal[
                "proposed_final"
            ]
        ).strip()

        if proposed:
            cell.value = (
                proposed
            )

            new_vocab = str(
                proposal[
                    "new_vocabulary"
                ]
            ).strip()

            if new_vocab:
                cell.fill = (
                    new_fill
                )

                cell.font = Font(
                    bold=True
                )

        else:
            cell.fill = (
                unresolved_fill
            )

            cell.font = Font(
                bold=True
            )

    # ======================================================
    # SUPPORTING SHEETS
    # ======================================================

    append_dataframe_sheet(
        workbook,
        "Classification Evidence V2.1",
        evidence_df,
    )

    append_dataframe_sheet(
        workbook,
        "Classification Review V2.1",
        unresolved_df,
    )

    append_dataframe_sheet(
        workbook,
        "New Vocabulary Usage V2.1",
        new_usage_df,
    )

    append_dataframe_sheet(
        workbook,
        "V2 vs V2.1",
        comparison_df,
    )

    # ======================================================
    # CLASSIFICATION LEGEND
    # ======================================================

    legend_name = (
        "Classification Legend V2.1"
    )

    if (
        legend_name
        in workbook.sheetnames
    ):
        del workbook[
            legend_name
        ]

    legend_ws = (
        workbook.create_sheet(
            legend_name
        )
    )

    legend_ws.append(
        [
            "Indicator",
            "Meaning",
        ]
    )

    legend_ws[
        "A1"
    ].font = Font(
        bold=True
    )

    legend_ws[
        "B1"
    ].font = Font(
        bold=True
    )

    legend_ws.append(
        [
            "Green cell in Klasifikasi Final",
            (
                "Final classification contains "
                "at least one NEW vocabulary "
                "supported by actual work evidence."
            ),
        ]
    )

    legend_ws.append(
        [
            "Purple cell in Klasifikasi Final",
            (
                "Classification could not be "
                "determined confidently. "
                "Manual review is required."
            ),
        ]
    )

    legend_ws.append(
        [
            "PO+ITEM_U",
            (
                "Classification supported by both "
                "raw PO and Item Pekerjaan "
                "Berdasarkan PO."
            ),
        ]
    )

    legend_ws.append(
        [
            "PO",
            (
                "Classification supported directly "
                "by raw PO description."
            ),
        ]
    )

    legend_ws.append(
        [
            "ITEM_U",
            (
                "Classification supported by "
                "Item Pekerjaan Berdasarkan PO."
            ),
        ]
    )

    legend_ws.append(
        [
            "TAXONOMY_FALLBACK",
            (
                "No primary rule matched, "
                "but a highly specific taxonomy "
                "classification was available."
            ),
        ]
    )

    legend_ws.append(
        [
            "CONFIRMED_CIRCLE",
            (
                "Actual work classification is also "
                "consistent with capability declared "
                "by vendor in HK Circle."
            ),
        ]
    )

    legend_ws.append(
        [
            "ACTUAL_WORK_NOT_DECLARED",
            (
                "Vendor performed the work, "
                "but the same capability was not "
                "declared in HK Circle."
            ),
        ]
    )

    legend_ws.column_dimensions[
        "A"
    ].width = 32

    legend_ws.column_dimensions[
        "B"
    ].width = 90

    workbook.save(
        output_file
    )

    # ======================================================
    # VERIFY OUTPUT
    # ======================================================

    verify = pd.read_excel(
        output_file,
        sheet_name=CLEANSING_SHEET,
        dtype=str,
        keep_default_na=False,
    )

    if (
        len(verify)
        != EXPECTED_VENDOR_ROWS
    ):
        raise RuntimeError(
            "OUTPUT ROW GUARD FAILED: "
            f"expected "
            f"{EXPECTED_VENDOR_ROWS:,}, "
            f"got "
            f"{len(verify):,}"
        )

    remaining_blank = int(
        verify[
            FINAL_COLUMN
        ]
        .astype(str)
        .str.strip()
        .eq("")
        .sum()
    )

    newly_classified = (
        target_count
        - len(
            unresolved_df
        )
    )

    print()
    print(
        "=" * 74
    )

    print(
        "CLASSIFICATION V2.1"
    )

    print(
        "=" * 74
    )

    print(
        f"Vendor rows             : "
        f"{len(cleansing):,}"
    )

    print(
        f"Existing final preserved: "
        f"{existing_count:,}"
    )

    print(
        f"Blank final target      : "
        f"{target_count:,}"
    )

    print(
        f"Newly classified        : "
        f"{newly_classified:,}"
    )

    print(
        f"Unresolved              : "
        f"{len(unresolved_df):,}"
    )

    print(
        f"Remaining blank T       : "
        f"{remaining_blank:,}"
    )

    print(
        f"New vocabulary used     : "
        f"{len(new_usage_df):,}"
    )

    print(
        f"Changed vs V2           : "
        f"{len(comparison_df):,}"
    )

    print()
    print(
        "Evidence source counts:"
    )

    for source in [
        "PO+ITEM_U",
        "PO",
        "ITEM_U",
        "TAXONOMY_FALLBACK",
    ]:
        print(
            f"  {source:<20}: "
            f"{source_counter[source]:,}"
        )

    print()
    print(
        f"Output                  : "
        f"{output_file}"
    )

    print(
        "=" * 74
    )


if __name__ == "__main__":
    main()