from __future__ import annotations

import shutil
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


# ============================================================
# PROJECT PATH
# ============================================================

ROOT = Path(__file__).resolve().parents[1]

if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


# ============================================================
# INTERNAL IMPORTS
# ============================================================

from src.audit import audit_cleansing
from src.config import (
    CLEANSING_FILE,
    CLEANSING_SHEET,
    EXPECTED_VENDOR_ROWS,
    OUTPUT_DIR,
    REQUIRED_CLEANSING_COLUMNS,
    ROW_FILL,
    SEVERITY,
)
from src.vocabulary import extract_uppercase_legacy_vocabulary


# ============================================================
# VALIDATION
# ============================================================


def require_columns(
    df: pd.DataFrame,
    required: list[str],
    source: str,
) -> None:
    missing = [
        column
        for column in required
        if column not in df.columns
    ]

    if missing:
        raise ValueError(
            f"{source}: missing columns: {missing}"
        )


# ============================================================
# MAIN
# ============================================================


def main() -> None:
    output_dir = OUTPUT_DIR / "01_audit"

    output_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    # --------------------------------------------------------
    # LOAD DATA CLEANSING ONLY
    # --------------------------------------------------------

    cleansing = pd.read_excel(
        CLEANSING_FILE,
        sheet_name=CLEANSING_SHEET,
        dtype=str,
        keep_default_na=False,
    )

    require_columns(
        cleansing,
        REQUIRED_CLEANSING_COLUMNS,
        "Data Cleansing",
    )

    # --------------------------------------------------------
    # INPUT ROW GUARD
    # --------------------------------------------------------

    if len(cleansing) != EXPECTED_VENDOR_ROWS:
        raise RuntimeError(
            "ROW GUARD FAILED: "
            f"expected {EXPECTED_VENDOR_ROWS:,} vendor rows, "
            f"got {len(cleansing):,}. "
            "Processing stopped."
        )

    # --------------------------------------------------------
    # AUDIT DATA CLEANSING
    # --------------------------------------------------------

    findings, row_summary = audit_cleansing(
        cleansing
    )

    # --------------------------------------------------------
    # EXTRACT LEGACY UPPERCASE VOCABULARY
    # --------------------------------------------------------

    vocabulary = extract_uppercase_legacy_vocabulary(
        cleansing
    )

    # --------------------------------------------------------
    # EXPORT CSV AUDIT
    # --------------------------------------------------------

    findings.to_csv(
        output_dir / "audit_findings.csv",
        index=False,
        encoding="utf-8-sig",
    )

    row_summary.to_csv(
        output_dir / "audit_row_summary.csv",
        index=False,
        encoding="utf-8-sig",
    )

    vocabulary.to_csv(
        output_dir / "vocabulary_v1.csv",
        index=False,
        encoding="utf-8-sig",
    )

    # --------------------------------------------------------
    # FINDING SUMMARY
    # --------------------------------------------------------

    if findings.empty:
        finding_counts = pd.DataFrame(
            columns=[
                "severity",
                "code",
                "row_findings",
            ]
        )

    else:
        finding_counts = (
            findings
            .groupby(
                [
                    "severity",
                    "code",
                ],
                as_index=False,
            )
            .size()
            .rename(
                columns={
                    "size": "row_findings",
                }
            )
        )

        finding_counts["severity_rank"] = (
            finding_counts["severity"]
            .map(SEVERITY)
        )

        finding_counts = (
            finding_counts
            .sort_values(
                [
                    "severity_rank",
                    "row_findings",
                    "code",
                ],
                ascending=[
                    False,
                    False,
                    True,
                ],
            )
            .drop(
                columns="severity_rank"
            )
        )

    finding_counts.to_csv(
        output_dir / "audit_finding_counts.csv",
        index=False,
        encoding="utf-8-sig",
    )

    # --------------------------------------------------------
    # CREATE REVIEW WORKBOOK
    # --------------------------------------------------------

    review_file = (
        output_dir
        / "Data Cleansing - Audit Review.xlsx"
    )

    shutil.copy2(
        CLEANSING_FILE,
        review_file,
    )

    workbook = load_workbook(
        review_file
    )

    worksheet = workbook[
        CLEANSING_SHEET
    ]

    # --------------------------------------------------------
    # HIGHLIGHT FLAGGED ROWS A:Y
    # --------------------------------------------------------

    if not row_summary.empty:
        severity_by_row = dict(
            zip(
                row_summary["excel_row"],
                row_summary["highest_severity"],
            )
        )

        for excel_row, severity in severity_by_row.items():
            fill = PatternFill(
                fill_type="solid",
                fgColor=ROW_FILL[severity],
            )

            # Data Cleansing has A:Y = 25 columns.
            for cell in worksheet[int(excel_row)][0:25]:
                cell.fill = fill

    # --------------------------------------------------------
    # CREATE AUDIT FINDINGS SHEET
    # --------------------------------------------------------

    if "Audit Findings" in workbook.sheetnames:
        del workbook["Audit Findings"]

    audit_ws = workbook.create_sheet(
        "Audit Findings"
    )

    audit_ws.append(
        [
            "Excel Row",
            "NO SAP",
            "Nama Rekanan",
            "Severity",
            "Flags",
            "Details",
        ]
    )

    for cell in audit_ws[1]:
        cell.font = Font(
            bold=True
        )

    cleansing_lookup = cleansing.copy()

    cleansing_lookup["__excel_row"] = range(
        2,
        len(cleansing_lookup) + 2,
    )

    cleansing_lookup = (
        cleansing_lookup
        .set_index("__excel_row")
    )

    sortable_summary = row_summary.copy()

    if not sortable_summary.empty:
        sortable_summary["__severity_rank"] = (
            sortable_summary[
                "highest_severity"
            ]
            .map(SEVERITY)
        )

        sortable_summary = (
            sortable_summary
            .sort_values(
                [
                    "__severity_rank",
                    "excel_row",
                ],
                ascending=[
                    False,
                    True,
                ],
            )
        )

        for _, finding in sortable_summary.iterrows():
            excel_row = int(
                finding["excel_row"]
            )

            source_row = cleansing_lookup.loc[
                excel_row
            ]

            audit_ws.append(
                [
                    excel_row,
                    source_row["NO SAP"],
                    source_row["Nama Rekanan"],
                    finding[
                        "highest_severity"
                    ],
                    finding["flags"],
                    finding["details"],
                ]
            )

    audit_ws.freeze_panes = "A2"

    audit_ws.auto_filter.ref = (
        audit_ws.dimensions
    )

    audit_ws.column_dimensions[
        "A"
    ].width = 12

    audit_ws.column_dimensions[
        "B"
    ].width = 18

    audit_ws.column_dimensions[
        "C"
    ].width = 40

    audit_ws.column_dimensions[
        "D"
    ].width = 14

    audit_ws.column_dimensions[
        "E"
    ].width = 55

    audit_ws.column_dimensions[
        "F"
    ].width = 100

    # --------------------------------------------------------
    # CREATE AUDIT LEGEND
    # --------------------------------------------------------

    if "Audit Legend" in workbook.sheetnames:
        del workbook["Audit Legend"]

    legend_ws = workbook.create_sheet(
        "Audit Legend"
    )

    legend_ws.append(
        [
            "Severity",
            "Meaning",
        ]
    )

    for cell in legend_ws[1]:
        cell.font = Font(
            bold=True
        )

    legend_rows = [
        (
            "CRITICAL",
            (
                "Hard key/data-integrity issue; "
                "row tetap dipertahankan tetapi wajib review."
            ),
        ),
        (
            "HIGH",
            (
                "Strong anomaly / possible duplicate; "
                "tidak boleh auto-merge."
            ),
        ),
        (
            "MEDIUM",
            (
                "Possible business duplicate berdasarkan "
                "nama/identity; row tetap dipertahankan."
            ),
        ),
        (
            "LOW",
            (
                "Classification source consistency issue "
                "seperti Level 2/3 kosong atau nama taxonomy "
                "tidak canonical."
            ),
        ),
    ]

    for severity, meaning in legend_rows:
        legend_ws.append(
            [
                severity,
                meaning,
            ]
        )

        current_row = (
            legend_ws.max_row
        )

        fill = PatternFill(
            fill_type="solid",
            fgColor=ROW_FILL[
                severity
            ],
        )

        legend_ws.cell(
            row=current_row,
            column=1,
        ).fill = fill

        legend_ws.cell(
            row=current_row,
            column=2,
        ).fill = fill

    legend_ws.column_dimensions[
        "A"
    ].width = 15

    legend_ws.column_dimensions[
        "B"
    ].width = 100

    # --------------------------------------------------------
    # SAVE
    # --------------------------------------------------------

    workbook.save(
        review_file
    )

    # --------------------------------------------------------
    # OUTPUT ROW GUARD
    # --------------------------------------------------------

    verify = pd.read_excel(
        review_file,
        sheet_name=CLEANSING_SHEET,
        dtype=str,
        keep_default_na=False,
    )

    if len(verify) != EXPECTED_VENDOR_ROWS:
        raise RuntimeError(
            "OUTPUT ROW GUARD FAILED: "
            f"expected {EXPECTED_VENDOR_ROWS:,}, "
            f"got {len(verify):,}."
        )

    # --------------------------------------------------------
    # SUMMARY
    # --------------------------------------------------------

    if (
        not vocabulary.empty
        and "enabled" in vocabulary.columns
    ):
        enabled_series = (
            vocabulary["enabled"]
            .astype(str)
            .str.upper()
            .isin(
                [
                    "TRUE",
                    "1",
                    "YES",
                ]
            )
        )

        enabled_vocab_count = int(
            enabled_series.sum()
        )

    else:
        enabled_vocab_count = len(
            vocabulary
        )

    print()
    print("=" * 70)
    print("STAGE 1 - DATA CLEANSING AUDIT")
    print("=" * 70)

    print(
        f"Vendor rows                : "
        f"{len(cleansing):,}"
    )

    print(
        f"Flagged vendor rows        : "
        f"{len(row_summary):,}"
    )

    print(
        f"Audit finding instances    : "
        f"{len(findings):,}"
    )

    print(
        f"Legacy vocabulary detected : "
        f"{len(vocabulary):,}"
    )

    print(
        f"Enabled vocabulary         : "
        f"{enabled_vocab_count:,}"
    )

    print(
        f"Output workbook            : "
        f"{review_file}"
    )

    print()
    print("Finding counts:")
    print("-" * 70)

    if finding_counts.empty:
        print("No findings.")

    else:
        print(
            finding_counts.to_string(
                index=False
            )
        )

    print("=" * 70)


if __name__ == "__main__":
    main()